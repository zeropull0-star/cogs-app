# api/main.py
# -*- coding: utf-8 -*-
from __future__ import annotations

import os
from contextlib import asynccontextmanager
from datetime import datetime, timedelta, timezone
from enum import Enum
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from fastapi import Depends, FastAPI, Form, HTTPException, Query, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from fastapi.security import OAuth2PasswordBearer
from jose import JWTError, jwt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from passlib.context import CryptContext
from pydantic import BaseModel, ConfigDict, Field
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from sqlalchemy import (Boolean, Column, DateTime, Enum as SAEnum, ForeignKey,
                        Index, Integer, Numeric, String, Text, create_engine, text)
from sqlalchemy.orm import Session, declarative_base, relationship, sessionmaker
from zoneinfo import ZoneInfo

# 나눔고딕 - 시스템 설치(apt fonts-nanum) 우선, assets 폴더 폴백
_NANUM_CANDIDATES = [
    ("/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
     "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf"),
    ("/app/assets/NanumGothic.ttf",
     "/app/assets/NanumGothicBold.ttf"),
]
_PDF_FONT = _PDF_FONT_BOLD = None
for _reg, _bold in _NANUM_CANDIDATES:
    try:
        pdfmetrics.registerFont(TTFont("NanumGothic", _reg))
        _PDF_FONT = "NanumGothic"
        try:
            pdfmetrics.registerFont(TTFont("NanumGothicBold", _bold))
            _PDF_FONT_BOLD = "NanumGothicBold"
        except Exception:
            _PDF_FONT_BOLD = "NanumGothic"
        break
    except Exception:
        continue
if not _PDF_FONT:
    _PDF_FONT = _PDF_FONT_BOLD = "Helvetica"

W, H = A4

# ─── Config ───────────────────────────────────────────────────
DATABASE_URL = os.getenv("DATABASE_URL", "postgresql+psycopg://postgres:postgres@db:5432/ledger")
SECRET_KEY   = os.getenv("SECRET_KEY", "CHANGE_ME_TO_A_RANDOM_LONG_SECRET").strip()
ALGORITHM    = (os.getenv("JWT_ALGORITHM", "HS256") or "HS256").strip()
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", "1440"))
ADMIN_USERNAME = (os.getenv("ADMIN_USERNAME", "admin") or "admin").strip()
ADMIN_PASSWORD = (os.getenv("ADMIN_PASSWORD", "admin1234") or "admin1234").strip()
DOCNO_PREFIX   = (os.getenv("DOCNO_PREFIX", "BWIS") or "BWIS").strip()
DOCNO_DATE_FMT = (os.getenv("DOCNO_DATE_FMT", "%Y%m%d") or "%Y%m%d").strip()

pwd_context   = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login")

engine       = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
Base         = declarative_base()

UTC = ZoneInfo("UTC")
KST = ZoneInfo("Asia/Seoul")


# ═══════════════ Enums ═════════════════════════════════════════
class TxKind(str, Enum):
    매입 = "매입"
    매출 = "매출"


class DocType(str, Enum):
    견적서    = "견적서"
    발주서    = "발주서"
    거래명세서 = "거래명세서"


# ═══════════════ ORM Models ════════════════════════════════════
class User(Base):
    __tablename__ = "users"
    id            = Column(Integer, primary_key=True)
    username      = Column(String(200), nullable=False, unique=True, index=True)
    password_hash = Column(String(300), nullable=False)
    is_active     = Column(Boolean, nullable=False, default=True)
    created_at    = Column(DateTime, default=lambda: datetime.now(timezone.utc), nullable=False)


class Company(Base):
    __tablename__ = "companies"
    id         = Column(Integer, primary_key=True)
    name       = Column(String(200), nullable=False)
    biz_no     = Column(String(64),  nullable=True)
    ceo        = Column(String(100), nullable=True)
    manager    = Column(String(100), nullable=True)
    addr       = Column(String(300), nullable=True)
    phone      = Column(String(50),  nullable=True)
    logo_path  = Column(String(300), nullable=True)
    seal_path  = Column(String(300), nullable=True)
    color      = Column(String(20),  nullable=True, default="#2563eb")
    doc_prefix = Column(String(16),  nullable=True)   # 문서번호 접두사
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc), nullable=False)


class Vendor(Base):
    __tablename__ = "vendors"
    id         = Column(Integer, primary_key=True)
    name       = Column(String(200), nullable=False, index=True)
    biz_no     = Column(String(64),  nullable=True)
    ceo        = Column(String(100), nullable=True)
    manager    = Column(String(100), nullable=True)
    addr       = Column(String(300), nullable=True)
    phone      = Column(String(50),  nullable=True)
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc), nullable=False)
    txs        = relationship("Tx", back_populates="vendor")


class Tx(Base):
    __tablename__ = "tx"
    id            = Column(Integer, primary_key=True)
    kind          = Column(SAEnum(TxKind, name="tx_kind"), nullable=False, index=True)
    tx_date       = Column(DateTime, nullable=False,
                           default=lambda: datetime.now(timezone.utc), index=True)
    vendor_id     = Column(Integer, ForeignKey("vendors.id", ondelete="RESTRICT"), nullable=False)
    company_id    = Column(Integer, ForeignKey("companies.id", ondelete="SET NULL"), nullable=True)
    description   = Column(Text, nullable=True)
    vat_rate      = Column(Numeric(6, 4),  nullable=False, default=0.10)
    supply_amount = Column(Numeric(18, 0), nullable=False, default=0)
    vat_amount    = Column(Numeric(18, 0), nullable=False, default=0)
    total_amount  = Column(Numeric(18, 0), nullable=False, default=0)
    doc_no        = Column(String(64), nullable=True, index=True)
    vendor  = relationship("Vendor", back_populates="txs")
    company = relationship("Company", foreign_keys=[company_id])
    items   = relationship("TxItem", back_populates="tx",
                          cascade="all, delete-orphan", order_by="TxItem.id")


class TxItem(Base):
    __tablename__ = "tx_items"
    id         = Column(Integer, primary_key=True)
    tx_id      = Column(Integer, ForeignKey("tx.id", ondelete="CASCADE"),
                        nullable=False, index=True)
    name       = Column(String(200), nullable=False)
    spec       = Column(String(200), nullable=True)
    qty        = Column(Numeric(18, 4), nullable=False, default=1)
    unit_price = Column(Numeric(18, 0), nullable=False, default=0)
    tx         = relationship("Tx", back_populates="items")


Index("ix_tx_kind_date", Tx.kind, Tx.tx_date.desc())


# ═══════════════ Auth helpers ══════════════════════════════════
def create_access_token(data: dict, expires_delta: Optional[timedelta] = None) -> str:
    to_encode = data.copy()
    expire    = datetime.now(timezone.utc) + (
        expires_delta or timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
    to_encode.update({"exp": expire, "iat": datetime.now(timezone.utc)})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def get_password_hash(pw: str) -> str:
    return pwd_context.hash(pw)


def db_session():
    return SessionLocal()


def init_db():
    with SessionLocal() as db:
        if not db.query(User).filter(User.username == ADMIN_USERNAME).first():
            db.add(User(username=ADMIN_USERNAME,
                        password_hash=get_password_hash(ADMIN_PASSWORD),
                        is_active=True))
            db.commit()
        if db.query(Company).count() == 0:
            db.add(Company(name="부원정보 주식회사", biz_no="328-81-02550",
                           ceo="홍길동", addr="강원특별자치도 동해시 중앙로 173 2층",
                           phone="010-4236-8448",
                           logo_path="/app/assets/logo.png",
                           seal_path="/app/assets/seal.png",
                           color="#2563eb",
                           doc_prefix="BWIS"))
            db.add(Company(name="동명정보 주식회사", biz_no="245-86-03526",
                           ceo="박정욱", manager=None,
                           addr="강원특별자치도 동해시 샘실2길 35(천곡동)",
                           phone="010-4236-8448",
                           color="#16a34a",
                           doc_prefix="DMIS"))
            db.add(Company(name="다한지엔에스(주)", biz_no="222-81-27364",
                           ceo="박민수", manager=None,
                           addr="강원특별자치도 삼척시 오십천로 102-12, 1층2호(등봉동)",
                           phone="010-4236-8448",
                           color="#dc2626",
                           doc_prefix="DHGNS"))
            db.commit()


def ensure_schema():
    # 1) ORM 테이블 생성 (신규 설치 시)
    Base.metadata.create_all(bind=engine)
    # 2) 기존 DB에 신규 컬럼 추가 (마이그레이션)
    try:
        with engine.begin() as conn:
            for sql in [
                "ALTER TABLE tx_items  ADD COLUMN IF NOT EXISTS spec    VARCHAR(200);",
                "ALTER TABLE companies ADD COLUMN IF NOT EXISTS manager VARCHAR(100);",
                "ALTER TABLE companies ADD COLUMN IF NOT EXISTS color   VARCHAR(20) DEFAULT '#2563eb';",
                "ALTER TABLE companies ADD COLUMN IF NOT EXISTS doc_prefix VARCHAR(16);",
                "ALTER TABLE vendors   ADD COLUMN IF NOT EXISTS manager VARCHAR(100);",
                "ALTER TABLE tx        ADD COLUMN IF NOT EXISTS company_id INTEGER REFERENCES companies(id) ON DELETE SET NULL;",
                # 기존 회사 데이터에 doc_prefix 설정 (이미 값 있으면 무시)
                "UPDATE companies SET doc_prefix='BWIS'  WHERE name='부원정보 주식회사' AND doc_prefix IS NULL;",
                "UPDATE companies SET doc_prefix='DMIS'  WHERE name='동명정보 주식회사' AND doc_prefix IS NULL;",
                "UPDATE companies SET doc_prefix='DHGNS' WHERE name='다한지엔에스(주)' AND doc_prefix IS NULL;",
            ]:
                conn.execute(text(sql))
    except Exception:
        pass


# ═══════════════ Schemas ═══════════════════════════════════════
class TokenOut(BaseModel):
    access_token: str; token_type: str = "bearer"; expires_in_min: int


class UserOut(BaseModel):
    id: int; username: str; is_active: bool


class CompanyIn(BaseModel):
    model_config = ConfigDict(extra="forbid")
    name: str; biz_no: Optional[str] = None; ceo: Optional[str] = None
    manager: Optional[str] = None
    addr: Optional[str] = None; phone: Optional[str] = None
    logo_path: Optional[str] = None; seal_path: Optional[str] = None
    color: Optional[str] = None
    doc_prefix: Optional[str] = None


class CompanyOut(CompanyIn):
    id: int; created_at: datetime


class VendorIn(BaseModel):
    model_config = ConfigDict(extra="forbid")
    name: str; biz_no: Optional[str] = None; ceo: Optional[str] = None
    manager: Optional[str] = None
    addr: Optional[str] = None; phone: Optional[str] = None


class VendorOut(VendorIn):
    id: int; created_at: datetime


class TxItemIn(BaseModel):
    model_config = ConfigDict(extra="forbid")
    name: str; spec: Optional[str] = None
    qty: float = Field(ge=0); unit_price: float = Field(ge=0)


class TxIn(BaseModel):
    model_config = ConfigDict(extra="forbid")
    kind: TxKind; tx_date: Optional[datetime] = None; vendor_id: int
    company_id: Optional[int] = None
    description: Optional[str] = None
    vat_rate: float = Field(default=0.10, ge=0, le=1)
    doc_no: Optional[str] = None
    items: List[TxItemIn] = Field(default_factory=list)


class TxOut(BaseModel):
    id: int; kind: TxKind; tx_date: datetime; vendor_id: int; vendor_name: str
    company_id: Optional[int]; description: Optional[str]; vat_rate: float
    supply_amount: int; vat_amount: int; total_amount: int
    doc_no: Optional[str]; items: List[TxItemIn]


# ═══════════════ Auth dep ══════════════════════════════════════
def get_current_user(token: str = Depends(oauth2_scheme)) -> User:
    exc = HTTPException(status_code=status.HTTP_401_UNAUTHORIZED,
                        detail="Unauthorized",
                        headers={"WWW-Authenticate": "Bearer"})
    try:
        payload  = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
        if not username:
            raise exc
    except JWTError:
        raise exc
    with db_session() as db:
        user = db.query(User).filter(User.username == username).first()
        if not user or not user.is_active:
            raise exc
        return user


# ═══════════════ Common helpers ════════════════════════════════
def _to_kst(dt: datetime) -> datetime:
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=UTC)
    return dt.astimezone(KST)


def _fmt_money(n: int) -> str:
    try:
        return f"{int(n):,}"
    except Exception:
        return "0"


def _safe_int(v) -> int:
    try:
        return int(round(float(v)))
    except Exception:
        return 0


def _gen_doc_no(dt: datetime, vendor_prefix: Optional[str] = None) -> str:
    prefix = (vendor_prefix or "").strip() or DOCNO_PREFIX
    return f"{prefix}-{dt.strftime(DOCNO_DATE_FMT)}"


def _calc_amounts(items: List[TxItemIn], vat_rate: float) -> Tuple[int, int, int]:
    supply = sum(_safe_int(it.qty * it.unit_price) for it in items)
    vat    = _safe_int(supply * float(vat_rate))
    return supply, vat, supply + vat


def _first_line(memo: Optional[str], max_len: int = 50) -> str:
    if not memo:
        return ""
    s    = memo.strip().replace("\r\n", "\n").replace("\r", "\n")
    line = s.split("\n")[0].strip()
    return line[: max_len - 1] + "…" if len(line) > max_len else line


def _tx_to_out(t: Tx, vendor_name: str, items: List[TxItem]) -> TxOut:
    return TxOut(
        id=t.id, kind=t.kind, tx_date=_to_kst(t.tx_date),
        vendor_id=t.vendor_id, vendor_name=vendor_name,
        company_id=t.company_id,
        description=t.description, vat_rate=float(t.vat_rate),
        supply_amount=_safe_int(t.supply_amount),
        vat_amount=_safe_int(t.vat_amount),
        total_amount=_safe_int(t.total_amount),
        doc_no=t.doc_no,
        items=[TxItemIn(name=i.name, spec=i.spec,
                        qty=float(i.qty), unit_price=float(i.unit_price))
               for i in items],
    )


# ═══════════════════════════════════════════════════════════════
#  PDF generation
# ═══════════════════════════════════════════════════════════════
_C_NAVY   = colors.HexColor("#1b2a4a")   # 주 색상 (네이비)
_C_BLUE   = colors.HexColor("#2563eb")   # 강조 (블루)
_C_ALT    = colors.HexColor("#f0f5ff")   # 짝수 행 배경
_C_BORDER = colors.HexColor("#cbd5e1")   # 일반 보더
_C_DARK_B = colors.HexColor("#64748b")   # 진한 보더
_C_TEXT   = colors.HexColor("#1e293b")   # 본문 텍스트
_C_MUTED  = colors.HexColor("#64748b")   # 보조 텍스트
_C_TOTAL  = colors.HexColor("#1b2a4a")   # 합계
_C_HDR_BG = colors.HexColor("#1b2a4a")   # 테이블 헤더 배경
_C_SUB_BG = colors.HexColor("#e8edf8")   # 소계 배경
# 하위 호환성
_C_TITLE  = _C_NAVY
FONT      = _PDF_FONT
M         = 15 * mm   # 여백 조금 줄여서 내용 공간 확보


def _maybe_img(c, path, x, y, w, h):
    if not path or not os.path.exists(path):
        return
    try:
        from reportlab.lib.utils import ImageReader
        c.drawImage(ImageReader(path), x, y, width=w, height=h,
                    mask="auto", preserveAspectRatio=True, anchor="c")
    except Exception:
        pass


# ── PDF 색상 ─────────────────────────────────────────────
def _pdf_color(hex6: str):
    h = hex6.lstrip("#")
    r = int(h[0:2],16)/255; g = int(h[2:4],16)/255; b = int(h[4:6],16)/255
    return colors.Color(r, g, b)


def _pdf_lighten(hex6: str, f: float):
    h = hex6.lstrip("#")
    r = int(h[0:2],16); g = int(h[2:4],16); b = int(h[4:6],16)
    r = min(255, int(r+(255-r)*f))
    g = min(255, int(g+(255-g)*f))
    b = min(255, int(b+(255-b)*f))
    return colors.Color(r/255, g/255, b/255)


# ── 공통 드로잉 헬퍼 ─────────────────────────────────────
def _filled_rect(c, x, y, w, h, fill, stroke_color=None, stroke_w=0.5):
    c.setFillColor(fill)
    if stroke_color:
        c.setStrokeColor(stroke_color); c.setLineWidth(stroke_w)
        c.rect(x, y, w, h, fill=1, stroke=1)
    else:
        c.rect(x, y, w, h, fill=1, stroke=0)


def _border_rect(c, x, y, w, h, color, lw=0.5):
    c.setStrokeColor(color); c.setLineWidth(lw)
    c.rect(x, y, w, h, fill=0, stroke=1)


def _hline(c, x1, x2, y, color, lw=0.4):
    c.setStrokeColor(color); c.setLineWidth(lw)
    c.line(x1, y, x2, y)


def _vline(c, x, y1, y2, color, lw=0.4):
    c.setStrokeColor(color); c.setLineWidth(lw)
    c.line(x, y1, x, y2)


def _text(c, x, y, txt, font, size, color, align="L"):
    c.setFont(font, size); c.setFillColor(color)
    t = str(txt) if txt is not None else ""
    if align == "R":   c.drawRightString(x, y, t)
    elif align == "C": c.drawCentredString(x, y, t)
    else:              c.drawString(x, y, t)


# ── 정보 2단 표 (공급받는자|공급자) ──────────────────────
# doc_type 이 '발주서' 인 경우:
#   우리(company)는 물건을 '공급받는자'(왼쪽), 상대(vendor)는 '공급자'(오른쪽)
# 견적서/거래명세서:
#   상대(vendor)가 '공급받는자'(왼쪽), 우리(company)가 '공급자(발행)'(오른쪽)
def _draw_info_grid(c, vendor, company, x, y, w, accent, light, xlight, border_c, doc_type=None):
    ROW = 7 * mm; LBL = 17 * mm; HALF = w / 2
    is_po = (getattr(doc_type, "value", None) == "발주서")
    left, right = (company, vendor) if is_po else (vendor, company)
    has_mgr = True  # 담당자 행 항상 표시 (값 없어도)
    rows = [
        ("상    호",   left.name    or "", right.name    or ""),
        ("대 표 자",   left.ceo     or "", right.ceo     or ""),
        ("사업자번호", left.biz_no  or "", right.biz_no  or ""),
        ("주    소",   left.addr    or "", right.addr    or ""),
    ]
    if has_mgr:
        rows.append(("담 당 자", left.manager or "", right.manager or ""))
    rows.append(("연 락 처", left.phone or "", right.phone or ""))

    n  = len(rows)
    bh = ROW * n
    top = y   # y = 박스 최상단, 헤더는 y에서 아래로 시작

    # 섹션 헤더 (y ~ y-HDR_H)
    HDR_H = 7 * mm
    _filled_rect(c, x,        top - HDR_H, HALF, HDR_H, accent)
    _filled_rect(c, x + HALF, top - HDR_H, HALF, HDR_H, accent)
    _vline(c, x+HALF, top - HDR_H, top, colors.HexColor("#ffffff"), 0.3)
    _text(c, x + HALF/2,        top - HDR_H + 2.5*mm, "공급받는자",    FONT, 8.5, colors.white, "C")
    _text(c, x + HALF + HALF/2, top - HDR_H + 2.5*mm, "공급자 (발행)", FONT, 8.5, colors.white, "C")
    top -= HDR_H

    # 데이터 행
    for i, (lb, lv, rv) in enumerate(rows):
        ry = top - ROW
        # 라벨 배경
        _filled_rect(c, x,        ry, LBL,       ROW, light)
        _filled_rect(c, x + HALF, ry, LBL,       ROW, xlight)
        # 구분선
        _hline(c, x, x+w, ry, border_c, 0.3)
        _vline(c, x+LBL,       ry, ry+ROW, border_c, 0.3)
        _vline(c, x+HALF,      ry, ry+ROW, border_c, 0.5)
        _vline(c, x+HALF+LBL, ry, ry+ROW, border_c, 0.3)
        # 텍스트
        mid = ry + ROW/2 - 1.5*mm
        _text(c, x + LBL/2,             mid, lb, FONT, 7.5, accent, "C")
        _text(c, x + LBL + 2*mm,        mid, lv, FONT, 8,   colors.HexColor("#1e293b"))
        _text(c, x+HALF + LBL/2,        mid, lb, FONT, 7.5, accent, "C")
        _text(c, x+HALF + LBL + 2*mm,   mid, rv, FONT, 8,   colors.HexColor("#1e293b"))
        top -= ROW

    # 외곽 테두리
    total_h = HDR_H + bh
    _border_rect(c, x, top, w, total_h, accent, 0.8)
    _vline(c, x+HALF, top, top + total_h, accent, 0.8)

    # 직인 - 발행회사(company) 데이터 영역 세로 중앙
    # 발주서면 우리(company)가 왼쪽(공급받는자) → 직인도 왼쪽 칸
    seal_path = company.seal_path if hasattr(company, "seal_path") else None
    seal_mm   = 18 * mm
    if seal_path and os.path.exists(seal_path):
        # top = 데이터 하단, top+bh = 데이터 상단
        data_center = top + bh / 2
        if is_po:
            seal_x = x + HALF - seal_mm - 1*mm   # 왼쪽 칸 우측 끝
        else:
            seal_x = x + w - seal_mm - 1*mm      # 전체 우측 끝
        _maybe_img(c, seal_path,
                   seal_x,
                   data_center - seal_mm / 2,
                   seal_mm, seal_mm)

    return top  # 박스 하단 y


# ── 메타 바 (일자/문서번호/구분/건명) ───────────────────
def _draw_meta(c, tx, x, y, w, accent, light):
    BAR = 7 * mm
    _filled_rect(c, x, y - BAR, w, BAR, accent)
    kst  = _to_kst(tx.tx_date).strftime("%Y년 %m월 %d일")
    segs = [("구    분", tx.kind.value),
            ("건    명", _first_line(tx.description, 40) or "")]
    sw = w / len(segs)
    for i, (lb, val) in enumerate(segs):
        cx = x + i*sw + sw/2
        _text(c, cx, y - BAR + 4.2*mm, lb,  FONT, 7,   colors.white, "C")
        _text(c, cx, y - BAR + 1.5*mm, val, FONT, 7.5, colors.HexColor("#bfdbfe"), "C")
        if i: _vline(c, x+i*sw, y-BAR, y, colors.HexColor("#3b5fa0"), 0.3)
    _border_rect(c, x, y-BAR, w, BAR, colors.HexColor("#1e3a8a"), 0.5)
    return y - BAR


# ── 텍스트 줄바꿈 헬퍼 ─────────────────────────────────
def _wrap_text(c, font, size, text, max_w):
    """주어진 너비에 맞게 텍스트를 줄바꿈. 줄 리스트 반환."""
    if not text: return [""]
    words = list(text)  # 한글은 글자 단위로 처리
    lines = []; cur = ""
    for ch in text:
        test = cur + ch
        if c.stringWidth(test, font, size) <= max_w:
            cur = test
        else:
            if cur: lines.append(cur)
            cur = ch
    if cur: lines.append(cur)
    return lines if lines else [""]


# ── 품목 테이블 (줄바꿈 지원) ────────────────────────────
def _draw_items(c, rows_data, col_widths, aligns, row_h, hdr_h,
                x, y, accent, alt_bg, border_c, wrap_cols=None):
    """wrap_cols: 줄바꿈 허용할 컬럼 인덱스 집합 (0-based)"""
    TW     = sum(col_widths)
    thead  = rows_data[0]
    trows  = rows_data[1:]
    PX     = 2.5*mm
    FS     = 8.5
    LINE_H = FS * 1.4  # 줄간격 (pt)

    # 각 데이터 행의 실제 높이 계산 (줄바꿈 고려)
    row_heights = []
    wrapped     = []   # [(col_idx, [lines])]
    for row in trows:
        max_lines = 1
        row_wrap  = []
        for ci, (cell, cw) in enumerate(zip(row, col_widths)):
            if wrap_cols and ci in wrap_cols:
                lines = _wrap_text(c, FONT, FS, str(cell or ""), cw - PX*2)
                row_wrap.append(lines)
                max_lines = max(max_lines, len(lines))
            else:
                row_wrap.append([str(cell or "")])
        actual_h = max(row_h, max_lines * LINE_H/72*25.4*mm + 3*mm)
        row_heights.append(actual_h)
        wrapped.append(row_wrap)

    total_h = hdr_h + sum(row_heights)

    # 헤더 배경
    _filled_rect(c, x, y-hdr_h, TW, hdr_h, accent)

    # 데이터 행 배경
    cy = y - hdr_h
    for i, rh in enumerate(row_heights):
        if i % 2 == 1:
            _filled_rect(c, x, cy-rh, TW, rh, alt_bg)
        cy -= rh

    # 수평선
    cy = y - hdr_h
    _hline(c, x, x+TW, y, border_c, 0.3)
    _hline(c, x, x+TW, cy, border_c, 0.3)
    for rh in row_heights:
        cy -= rh
        _hline(c, x, x+TW, cy, border_c, 0.3)

    # 수직선
    cx2 = x
    for cw in col_widths[:-1]:
        cx2 += cw
        _vline(c, cx2, y, y-total_h, border_c, 0.3)

    # 외곽
    _border_rect(c, x, y-total_h, TW, total_h, accent, 0.7)

    # 헤더 텍스트
    ty = y - hdr_h + 2.8*mm
    c.setFillColor(colors.white); c.setFont(FONT, 9)
    cur_x = x
    for cell, cw, al in zip(thead, col_widths, aligns):
        t = str(cell or "")
        if al == "R":   c.drawRightString(cur_x+cw-PX, ty, t)
        elif al == "C": c.drawCentredString(cur_x+cw/2, ty, t)
        else:           c.drawString(cur_x+PX, ty, t)
        cur_x += cw

    # 데이터 텍스트
    cy = y - hdr_h
    for i, (row_wrap, rh) in enumerate(zip(wrapped, row_heights)):
        fc = colors.HexColor("#1e293b")
        c.setFillColor(fc); c.setFont(FONT, FS)
        cur_x = x
        for ci, (lines, cw, al) in enumerate(zip(row_wrap, col_widths, aligns)):
            n_lines = len(lines)
            # 여러 줄이면 세로 중앙 정렬
            total_text_h = n_lines * LINE_H/72*25.4*mm
            start_y = cy - rh/2 + total_text_h/2 - LINE_H/72*25.4*mm*0.3
            for li, line in enumerate(lines):
                ty2 = start_y - li * LINE_H/72*25.4*mm
                if al == "R":   c.drawRightString(cur_x+cw-PX, ty2, line)
                elif al == "C": c.drawCentredString(cur_x+cw/2, ty2, line)
                else:           c.drawString(cur_x+PX, ty2, line)
            cur_x += cw
        cy -= rh

    return total_h


# ── 합계 박스 ────────────────────────────────────────────
def _draw_summary_box(c, tx, x, y, tw, accent, light, xlight, border_c):
    SW   = 78 * mm; ROW = 8.5 * mm; TOT = 11 * mm
    bx   = x + tw - SW
    rows = [
        ("공  급  가  액", _fmt_money(_safe_int(tx.supply_amount)) + " 원"),
        ("부    가    세", _fmt_money(_safe_int(tx.vat_amount))    + " 원"),
    ]
    sh = ROW * len(rows) + TOT
    sy = y - sh

    # 소계 행
    for i, (lb, val) in enumerate(rows):
        ry = sy + sh - (i+1)*ROW
        _filled_rect(c, bx,         ry, 26*mm, ROW, light)
        _filled_rect(c, bx+26*mm,   ry, SW-26*mm, ROW, xlight)
        _hline(c, bx, bx+SW, ry, border_c, 0.3)
        _vline(c, bx+26*mm, ry, ry+ROW, border_c, 0.3)
        _text(c, bx+13*mm,   ry+ROW/2-1.5*mm, lb,  FONT, 8,  accent, "C")
        _text(c, bx+SW-3*mm, ry+ROW/2-1.5*mm, val, FONT, 8.5, colors.HexColor("#1e293b"), "R")

    # 합계 행
    _filled_rect(c, bx, sy, SW, TOT, accent)
    _text(c, bx+4*mm,   sy+3.5*mm, "합    계", FONT, 10, colors.white)
    _text(c, bx+SW-3*mm, sy+3.5*mm,
          _fmt_money(_safe_int(tx.total_amount)) + " 원",
          FONT, 11, colors.white, "R")

    # 외곽
    _border_rect(c, bx, sy, SW, sh, accent, 0.7)
    _hline(c, bx, bx+SW, sy+TOT, border_c, 0.3)


# ── 첫 페이지 헤더 ───────────────────────────────────────
def _draw_first_header(c, doc_type, tx, vendor, company):
    TW     = W - 2*M
    accent = _pdf_color(company.color or "#2563eb") if hasattr(company,"color") else _pdf_color("#2563eb")
    light  = _pdf_lighten(company.color or "#2563eb", 0.85)
    xlight = _pdf_lighten(company.color or "#2563eb", 0.93)
    border_c = _pdf_lighten(company.color or "#2563eb", 0.50)

    # ── 제목 바 ──
    TITLE_H = 13 * mm
    _filled_rect(c, M, H-M-TITLE_H, TW, TITLE_H, accent)
    _filled_rect(c, M, H-M-TITLE_H, 4*mm, TITLE_H, _pdf_lighten(company.color or "#2563eb", 0.15))
    title = "  ".join(doc_type.value)
    _text(c, W/2, H-M-TITLE_H+4*mm, title, FONT, 17, colors.white, "C")

    # 로고
    logo = company.logo_path if company.logo_path else "/app/assets/logo.png"
    _maybe_img(c, logo, M+6*mm, H-M-TITLE_H+1*mm, 38*mm, 11*mm)

    # ── 문서번호/구분/날짜 바 (제목 바로 아래, 한 줄) ──
    DOC_H = 8 * mm
    doc_y = H - M - TITLE_H - DOC_H
    _filled_rect(c, M, doc_y, TW, DOC_H, light)
    _border_rect(c, M, doc_y, TW, DOC_H, accent, 0.5)
    kst_doc = _to_kst(tx.tx_date).strftime("%Y년 %m월 %d일")
    ty = doc_y + 2.5*mm   # 텍스트 베이스라인 (아래서 위로)
    # 문서번호
    _text(c, M+4*mm,          ty, "문서번호", FONT, 7.5, accent)
    _text(c, M+22*mm,         ty, tx.doc_no or "", FONT, 9, colors.HexColor("#1e293b"))
    # 구분선
    _vline(c, M+TW*0.45, doc_y+1.5*mm, doc_y+DOC_H-1.5*mm, accent, 0.3)
    # 구분
    _text(c, M+TW*0.45+4*mm,  ty, "구  분", FONT, 7.5, accent)
    _text(c, M+TW*0.45+16*mm, ty, tx.kind.value, FONT, 9, colors.HexColor("#1e293b"))
    # 날짜 (우측)
    _text(c, W-M-4*mm,        ty, kst_doc, FONT, 8.5, colors.HexColor("#1e293b"), "R")

    # ── 정보 그리드 (공급받는자/공급자) ──
    grid_y      = doc_y - 2*mm
    grid_bottom = _draw_info_grid(c, vendor, company, M, grid_y, TW,
                                  accent, light, xlight, border_c, doc_type)

    # ── 건명 바 (정보 그리드 아래, 크게) ──
    memo_txt = _first_line(tx.description, 60) or ""
    if memo_txt:
        MEMO_H = 8 * mm
        memo_y = grid_bottom - 2*mm - MEMO_H
        _filled_rect(c, M, memo_y, TW, MEMO_H, xlight)
        _border_rect(c, M, memo_y, TW, MEMO_H, accent, 0.5)
        _text(c, M+4*mm,  memo_y+2.2*mm, "건    명", FONT, 8, accent)
        _text(c, M+22*mm, memo_y+2.2*mm, memo_txt, FONT, 10, colors.HexColor("#1e293b"))
        return memo_y - 3*mm
    return grid_bottom - 3*mm


# ── 이어지는 페이지 헤더 ─────────────────────────────────
def _draw_cont_header(c, doc_type, tx, page, company):
    TW     = W - 2*M
    accent = _pdf_color(company.color or "#2563eb") if hasattr(company,"color") else _pdf_color("#2563eb")
    BH = 9*mm
    _filled_rect(c, M, H-M-BH, TW, BH, accent)
    _filled_rect(c, M, H-M-BH, 4*mm, BH, _pdf_lighten(company.color or "#2563eb", 0.15))
    label = f"{doc_type.value}  ·  {tx.doc_no or ''}  ·  {page}페이지"
    _text(c, M+8*mm, H-M-BH+3*mm, label, FONT, 9, colors.white)
    return H - M - BH - 3*mm


# ── 메인 PDF 빌더 ────────────────────────────────────────
def build_pdf(doc_type: DocType, tx: Tx, vendor: Vendor,
              company: Company, items: List[TxItem]) -> bytes:
    buf    = BytesIO()
    c      = canvas.Canvas(buf, pagesize=A4)
    TW     = W - 2*M
    ROW_H  = 7.5*mm; HDR_H = 9.5*mm
    SUM_H  = 25*mm;  GAP   = 4*mm
    COL_W  = [9*mm, round(TW*0.20,1), round(TW*0.31,1), 14*mm,
              round(TW*0.15,1), round(TW*0.15,1)]
    COL_W[-1] = TW - sum(COL_W[:-1])
    ALIGNS    = ["C","L","L","C","R","R"]
    WRAP_COLS = {1, 2}   # 품목명, 규격 줄바꿈 허용

    accent   = _pdf_color(company.color or "#2563eb") if hasattr(company,"color") else _pdf_color("#2563eb")
    light    = _pdf_lighten(company.color or "#2563eb", 0.85)
    xlight   = _pdf_lighten(company.color or "#2563eb", 0.93)
    border_c = _pdf_lighten(company.color or "#2563eb", 0.55)

    total = len(items); page = 1; idx = 0

    while True:
        if page == 1:
            table_top = _draw_first_header(c, doc_type, tx, vendor, company)
        else:
            table_top = _draw_cont_header(c, doc_type, tx, page, company)

        avail_ws  = table_top - (M + SUM_H + GAP) - HDR_H
        avail_wos = table_top - M - HDR_H - GAP
        max_ws    = max(0, int(avail_ws  / ROW_H))
        max_wos   = max(0, int(avail_wos / ROW_H))

        remaining = items[idx:]
        if len(remaining) <= max_ws:
            items_this, is_last = remaining, True
        else:
            items_this = remaining[:max_wos]
            is_last    = (idx + len(items_this) >= total)

        idx += len(items_this); base_num = idx - len(items_this) + 1

        thead = ["번호", "품  목  명", "규    격", "수량", "단 가 (원)", "금 액 (원)"]
        trows = [[
            str(base_num+j), it.name or "", it.spec or "",
            f"{float(it.qty):g}",
            _fmt_money(_safe_int(it.unit_price)),
            _fmt_money(_safe_int(float(it.qty)*float(it.unit_price))),
        ] for j, it in enumerate(items_this)]

        tbl_h = _draw_items(c, [thead]+trows, COL_W, ALIGNS, ROW_H, HDR_H,
                    M, table_top, accent, xlight, border_c, WRAP_COLS)

        if is_last:
            _draw_summary_box(c, tx, M, table_top-tbl_h-GAP, TW,
                              accent, light, xlight, border_c)
        else:
            _text(c, W-M, table_top-tbl_h-3*mm,
                  f"▼ 다음 페이지에 계속  ({idx}/{total}건)",
                  FONT, 7.5, colors.HexColor("#64748b"), "R")

        _text(c, W/2, 10*mm, f"— {page} —", FONT, 7.5, colors.HexColor("#64748b"), "C")

        if is_last: break
        c.showPage(); page += 1

    c.save()
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════
#  Excel document generation
# ═══════════════════════════════════════════════════════════════
def _fill(hex_: str) -> PatternFill:
    h = hex_.lstrip("#")
    if len(h) == 6: h = "FF" + h   # ARGB: alpha=FF (불투명)
    return PatternFill("solid", fgColor=h)


def _outer() -> Border:
    """외곽 굵은 테두리"""
    s = Side(border_style="medium", color="1b2a4a")
    return Border(left=s, right=s, top=s, bottom=s)


def _inner_h() -> Border:
    """상하 얇은선"""
    th = Side(border_style="thin", color="94a3b8")
    return Border(top=th, bottom=th)


def _cell_border(top=False, bottom=False, left=False, right=False,
                 outer=False) -> Border:
    """세밀한 테두리 조합"""
    med = Side(border_style="medium", color="1b2a4a")
    thn = Side(border_style="thin",   color="94a3b8")
    non = Side(border_style=None)
    def _s(flag, med_flag=False): return med if med_flag else (thn if flag else non)
    return Border(
        top=   _s(top or outer,    outer),
        bottom=_s(bottom or outer, outer),
        left=  _s(left or outer,   outer),
        right= _s(right or outer,  outer),
    )


def build_excel(doc_type: DocType, tx: Tx, vendor: Vendor,
                company: Company, items: List[TxItem]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = doc_type.value

    # ── 인쇄 설정 ─────────────────────────────────────────────
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.5; ws.page_margins.right  = 0.5
    ws.page_margins.top  = 0.6; ws.page_margins.bottom = 0.6

    # ── 8열 구조 (완성본 기준) ────────────────────────────────
    # A=번호/라벨, B=품목명/라벨확장, C=규격/값, D=규격병합,
    # E=수량(좁음), F=단가/공급자라벨, G=금액/공급자값, H=비고/공급자값
    for col, w in zip("ABCDEFGH", [4.1, 23.8, 13, 13, 5.1, 27.9, 17.6, 16.6]):
        ws.column_dimensions[col].width = w

    # ── 회사 색상 ─────────────────────────────────────────────
    raw_color = (company.color or "#2563eb").lstrip("#")
    if len(raw_color) != 6: raw_color = "2563eb"

    def _lighten(hex6, factor):
        r=int(hex6[0:2],16); g=int(hex6[2:4],16); b=int(hex6[4:6],16)
        r=min(255,int(r+(255-r)*factor)); g=min(255,int(g+(255-g)*factor)); b=min(255,int(b+(255-b)*factor))
        return f"FF{r:02X}{g:02X}{b:02X}"

    C_HDR  = _fill(raw_color)                   # 제목/헤더 (원색)
    C_LBL  = _fill(_lighten(raw_color, 0.93))   # 왼쪽 라벨
    C_LBL2 = _fill(_lighten(raw_color, 0.87))   # 오른쪽 라벨
    C_ALT  = _fill(_lighten(raw_color, 0.93))   # 짝수행
    C_WHT  = _fill("FFFFFF")
    MED_C  = raw_color
    THN_C  = _lighten(raw_color, 0.50)[2:]

    FN = "맑은 고딕"

    def mc(r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    def wr(r, c, v="", bold=False, size=10, color="1E293B", bg=None,
           align="left", border=None, numfmt=None, wrap=False):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font      = Font(bold=bold, size=size, color=color, name=FN)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if bg is not None: cell.fill   = bg
        if border:         cell.border = border
        if numfmt:         cell.number_format = numfmt
        return cell

    def box():
        s = Side(border_style="thin", color="000000")
        return Border(left=s, right=s, top=s, bottom=s)

    R = 1
    kst_date = _to_kst(tx.tx_date).strftime("%Y년 %m월 %d일")

    # ══ 1. 제목 (A:H 전체) ════════════════════════════════════
    mc(R, 1, R, 8)
    wr(R, 1, "  ".join(doc_type.value), bold=True, size=28, color="FFFFFF",
       bg=C_HDR, align="center")
    ws.row_dimensions[R].height = 38
    R += 1

    # ══ 2. 문서번호 바 ════════════════════════════════════════
    # A:B=라벨, C:E=값, F=구분라벨, G=값, H=날짜
    ws.row_dimensions[R].height = 30
    mc(R, 1, R, 2)
    wr(R, 1, "문서번호", bold=True, size=10, color=raw_color, bg=C_LBL, align="center")
    mc(R, 3, R, 5)
    wr(R, 3, tx.doc_no or "—", size=11, bg=C_WHT, align="center")
    wr(R, 6, "구  분", bold=True, size=10, color=raw_color, bg=C_LBL, align="center")
    wr(R, 7, tx.kind.value, size=10, bg=C_WHT, align="center")
    wr(R, 8, kst_date, size=10, color=raw_color, bg=C_LBL, align="right")
    R += 1

    # ══ 3. 공급받는자 | 공급자 헤더 ══════════════════════════
    # A:E=공급받는자, F:H=공급자
    ws.row_dimensions[R].height = 35
    mc(R, 1, R, 5)
    wr(R, 1, "공  급  받  는  자", bold=True, size=10, color="FFFFFF", bg=C_HDR, align="center")
    mc(R, 6, R, 8)
    wr(R, 6, "공  급  자  (발  행)", bold=True, size=10, color="FFFFFF", bg=C_HDR, align="center")
    R += 1

    # ══ 4. 정보 그리드 ════════════════════════════════════════
    # 왼쪽(공급받는자): A:B=라벨, C:E=값 / 오른쪽(공급자): F=라벨, G:H=값
    # 발주서는 우리(company)가 공급받는자(왼쪽) ↔ 상대(vendor)가 공급자(오른쪽)
    is_po = (doc_type == DocType.발주서)
    left_src, right_src = (company, vendor) if is_po else (vendor, company)
    has_manager = True
    left_rows  = [("상    호", left_src.name or ""), ("대 표 자", left_src.ceo or ""),
                  ("사업자번호", left_src.biz_no or ""), ("주    소", left_src.addr or ""),
                  ("담 당 자", left_src.manager or ""), ("연 락 처", left_src.phone or "")]
    right_rows = [("상    호", right_src.name or ""), ("대 표 자", right_src.ceo or ""),
                  ("사업자번호", right_src.biz_no or ""), ("주    소", right_src.addr or ""),
                  ("담 당 자", right_src.manager or ""), ("연 락 처", right_src.phone or "")]

    seal_row = R
    for i, ((ll, lv), (rl, rv)) in enumerate(zip(left_rows, right_rows)):
        ws.row_dimensions[R].height = 22
        mc(R, 1, R, 2); wr(R, 1, ll, bold=True, size=10, color=raw_color, bg=C_LBL, align="center")
        mc(R, 3, R, 5); wr(R, 3, lv, size=10, bg=C_WHT, align="left")
        wr(R, 6, rl, bold=True, size=10, color=raw_color, bg=C_LBL2, align="center")
        if i == 0:
            wr(R, 7, rv, size=10, bg=C_WHT, align="left")
            wr(R, 8, "", bg=C_WHT)
        else:
            mc(R, 7, R, 8); wr(R, 7, rv, size=10, bg=C_WHT, align="left")
        R += 1

    # 직인 - 발행회사(company) 위치에 찍음
    # 발주서면 우리(company)가 왼쪽(공급받는자) → 왼쪽 영역(E열)
    SEAL_PX = 87
    if company.seal_path and os.path.exists(company.seal_path):
        try:
            img = XLImage(company.seal_path)
            img.width = SEAL_PX; img.height = SEAL_PX
            img.anchor = f"E{seal_row}" if is_po else f"H{seal_row}"
            ws.add_image(img)
        except Exception: pass

    # ══ 5. 건명 (A:B=라벨 3행병합, C:H=값 3행병합) ══════════
    mc(R, 1, R+2, 2)
    wr(R, 1, "건    명", bold=True, size=8, color=raw_color, bg=C_LBL, align="center")
    mc(R, 3, R+2, 8)
    memo = _first_line(tx.description, 60) or ""
    wr(R, 3, memo, size=14, bg=C_WHT, align="center")
    ws.row_dimensions[R].height = 6
    ws.row_dimensions[R+1].height = 26
    ws.row_dimensions[R+2].height = 6
    R += 3

    # ══ 6. 품목 헤더 ══════════════════════════════════════════
    # A=번호, B=품목명, C:D=규격, E=수량, F=단가, G=금액, H=비고
    ws.row_dimensions[R].height = 24
    mc(R, 3, R, 4)  # 규격 C:D 병합
    for ci, hd in [(1,"번호"),(2,"품  목  명"),(3,"규    격"),
                   (5,"수량"),(6,"단 가 (원)"),(7,"금 액 (원)"),(8,"비 고")]:
        wr(R, ci, hd, bold=True, size=9.5, color="FFFFFF", bg=C_HDR, align="center")
    ws.freeze_panes = f"A{R+1}"
    R += 1

    # ══ 7. 품목 데이터 ════════════════════════════════════════
    item_start = R
    for j, it in enumerate(items, 1):
        amt = _safe_int(float(it.qty) * float(it.unit_price))
        bg  = C_ALT if j % 2 == 0 else C_WHT
        name_lines = max(1, len(str(it.name or "")) // 20 + 1)
        spec_lines = max(1, len(str(it.spec or "")) // 13 + 1)
        ws.row_dimensions[R].height = max(17, max(name_lines, spec_lines) * 17)
        mc(R, 3, R, 4)
        wr(R, 1, j,                        size=10, bg=bg, align="center")
        wr(R, 2, it.name or "",            size=10, bg=bg, align="left",   wrap=True)
        wr(R, 3, it.spec or "",            size=10, bg=bg, align="center", wrap=True)
        wr(R, 5, float(it.qty),            size=10, bg=bg, align="center", numfmt="0.##")
        wr(R, 6, _safe_int(it.unit_price), size=10, bg=bg, align="right",  numfmt="#,##0")
        wr(R, 7, amt,                      size=10, bg=bg, align="right",   numfmt="#,##0")
        wr(R, 8, "",                       size=10, bg=bg, align="left")
        R += 1

    # 빈행 패딩 (A4 기준)
    FIXED = (item_start - 1) + 4
    padded = max(5, 56 - FIXED - len(items))
    for pi in range(padded):
        ws.row_dimensions[R].height = 17
        mc(R, 3, R, 4)
        for ci in range(1, 9):
            ws.cell(row=R, column=ci).fill = C_ALT if (len(items)+pi+1) % 2 == 0 else C_WHT
        R += 1

    # 여백
    ws.row_dimensions[R].height = 6; R += 1

    # ══ 8. 합계 ══════════════════════════════════════════════
    # A:E=라벨, F:H=값
    for lb, val, is_total in [
        ("공  급  가  액", _safe_int(tx.supply_amount), False),
        ("부    가    세", _safe_int(tx.vat_amount),    False),
        ("합        계",   _safe_int(tx.total_amount),  True),
    ]:
        ws.row_dimensions[R].height = 24 if is_total else 18
        bg_l = C_HDR if is_total else C_LBL
        bg_v = C_HDR if is_total else C_LBL
        fc   = "FFFFFF" if is_total else raw_color
        sz   = 12 if is_total else 9
        mc(R, 1, R, 5); wr(R, 1, lb, bold=is_total, size=sz, color=fc, bg=bg_l, align="center")
        mc(R, 6, R, 8); wr(R, 6, val, bold=is_total, size=sz, color=fc, bg=bg_v, align="right", numfmt="#,##0")
        R += 1

    # ══ 9. 전체 테두리 ════════════════════════════════════════
    tk = Side(border_style="thin", color="000000")
    fb = Border(top=tk, bottom=tk, left=tk, right=tk)
    for ri in range(1, R):
        for ci in range(1, 9):
            ws.cell(row=ri, column=ci).border = fb

    ws.print_area = f"A1:H{R-1}"
    bio = BytesIO(); wb.save(bio)
    return bio.getvalue()




# ═══════════════════════════════════════════════════════════════
#  FastAPI app
# ═══════════════════════════════════════════════════════════════
@asynccontextmanager
async def lifespan(_: FastAPI):
    ensure_schema()   # ← 먼저: 기존 테이블에 신규 컬럼 추가
    init_db()         # ← 이후: 테이블 생성 및 초기 데이터
    yield


app = FastAPI(title="Ledger API", lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"],
                   allow_credentials=False, allow_headers=["*"], allow_methods=["*"])


@app.get("/health")
def health():
    return {"ok": True}


# ── Auth ──────────────────────────────────────────────────────
@app.post("/api/auth/login", response_model=TokenOut)
def login(username: str = Form(...), password: str = Form(...)):
    with db_session() as db:
        u = db.query(User).filter(User.username == username).first()
        if not u or not verify_password(password, u.password_hash) or not u.is_active:
            raise HTTPException(400, "Invalid credentials")
        return TokenOut(access_token=create_access_token({"sub": u.username}),
                        expires_in_min=ACCESS_TOKEN_EXPIRE_MINUTES)


@app.get("/api/auth/me", response_model=UserOut)
def me(user: User = Depends(get_current_user)):
    return UserOut(id=user.id, username=user.username, is_active=user.is_active)


# ── Companies ─────────────────────────────────────────────────
def _co_out(co: Company) -> CompanyOut:
    return CompanyOut(id=co.id, name=co.name, biz_no=co.biz_no, ceo=co.ceo,
                      manager=co.manager, addr=co.addr, phone=co.phone,
                      logo_path=co.logo_path, seal_path=co.seal_path,
                      color=co.color, doc_prefix=co.doc_prefix,
                      created_at=co.created_at)


@app.get("/api/companies", response_model=List[CompanyOut])
def list_companies(user: User = Depends(get_current_user)):
    with db_session() as db:
        return [_co_out(co) for co in db.query(Company).order_by(Company.id).all()]


@app.post("/api/companies", response_model=CompanyOut)
def create_company(payload: CompanyIn, user: User = Depends(get_current_user)):
    with db_session() as db:
        co = Company(**payload.model_dump()); db.add(co); db.commit(); db.refresh(co)
        return _co_out(co)


@app.put("/api/companies/{co_id}", response_model=CompanyOut)
def update_company(co_id: int, payload: CompanyIn, user: User = Depends(get_current_user)):
    with db_session() as db:
        co = db.query(Company).filter(Company.id == co_id).first()
        if not co: raise HTTPException(404, "Company not found")
        for k, v in payload.model_dump().items(): setattr(co, k, v)
        db.commit(); db.refresh(co)
        return _co_out(co)


@app.delete("/api/companies/{co_id}")
def delete_company(co_id: int, user: User = Depends(get_current_user)):
    with db_session() as db:
        co = db.query(Company).filter(Company.id == co_id).first()
        if not co: raise HTTPException(404, "Company not found")
        db.delete(co); db.commit()
        return {"ok": True}


# ── Vendors ───────────────────────────────────────────────────
def _vnd_out(v: Vendor) -> VendorOut:
    return VendorOut(id=v.id, name=v.name, biz_no=v.biz_no, ceo=v.ceo,
                     manager=v.manager, addr=v.addr, phone=v.phone,
                     created_at=v.created_at)


@app.get("/api/vendors", response_model=List[VendorOut])
def list_vendors(q: Optional[str] = Query(default=None),
                 user: User = Depends(get_current_user)):
    with db_session() as db:
        qs = db.query(Vendor)
        if q: qs = qs.filter(Vendor.name.ilike(f"%{q}%"))
        return [_vnd_out(v) for v in qs.order_by(Vendor.id.desc()).all()]


@app.post("/api/vendors", response_model=VendorOut)
def create_vendor(payload: VendorIn, user: User = Depends(get_current_user)):
    with db_session() as db:
        v = Vendor(**payload.model_dump()); db.add(v); db.commit(); db.refresh(v)
        return _vnd_out(v)


@app.put("/api/vendors/{vid}", response_model=VendorOut)
def update_vendor(vid: int, payload: VendorIn, user: User = Depends(get_current_user)):
    with db_session() as db:
        v = db.query(Vendor).filter(Vendor.id == vid).first()
        if not v: raise HTTPException(404, "Vendor not found")
        for k, val in payload.model_dump().items(): setattr(v, k, val)
        db.commit(); db.refresh(v)
        return _vnd_out(v)


@app.delete("/api/vendors/{vid}")
def delete_vendor(vid: int, user: User = Depends(get_current_user)):
    with db_session() as db:
        v = db.query(Vendor).filter(Vendor.id == vid).first()
        if not v: raise HTTPException(404, "Vendor not found")
        db.delete(v); db.commit()
        return {"ok": True}


# ── TX ────────────────────────────────────────────────────────
def _items_of(db: Session, tx_id: int) -> List[TxItem]:
    return db.query(TxItem).filter(TxItem.tx_id == tx_id).order_by(TxItem.id).all()


@app.post("/api/tx", response_model=TxOut)
def create_tx(payload: TxIn, user: User = Depends(get_current_user)):
    tx_date          = payload.tx_date or datetime.now(timezone.utc)
    supply, vat, tot = _calc_amounts(payload.items, payload.vat_rate)
    with db_session() as db:
        vendor = db.query(Vendor).filter(Vendor.id == payload.vendor_id).first()
        if not vendor: raise HTTPException(400, "Invalid vendor_id")
        company = db.query(Company).filter(Company.id == payload.company_id).first() if payload.company_id else None
        t = Tx(kind=payload.kind, tx_date=tx_date, vendor_id=payload.vendor_id,
               company_id=payload.company_id,
               description=payload.description, vat_rate=payload.vat_rate,
               supply_amount=supply, vat_amount=vat, total_amount=tot,
               doc_no=(payload.doc_no or "").strip() or None)
        db.add(t); db.flush()
        if not t.doc_no:
            prefix = (company.doc_prefix or "").strip() if company else None
            t.doc_no = f"{_gen_doc_no(tx_date, prefix)}-{t.id:04d}"
        for it in payload.items:
            db.add(TxItem(tx_id=t.id, name=it.name,
                          spec=(it.spec or "").strip() or None,
                          qty=it.qty, unit_price=it.unit_price))
        db.commit(); db.refresh(t)
        return _tx_to_out(t, vendor.name, _items_of(db, t.id))


@app.get("/api/tx", response_model=List[TxOut])
def list_tx(kind: Optional[TxKind] = Query(default=None),
            vendor_id: Optional[int] = Query(default=None),
            limit: int = Query(default=200, ge=1, le=500),
            user: User = Depends(get_current_user)):
    with db_session() as db:
        qs = db.query(Tx).order_by(Tx.id.desc())
        if kind:      qs = qs.filter(Tx.kind == kind)
        if vendor_id: qs = qs.filter(Tx.vendor_id == vendor_id)
        txs     = qs.limit(limit).all()
        tx_ids  = [t.id for t in txs]
        vnd_ids = list({t.vendor_id for t in txs})
        all_its = (db.query(TxItem).filter(TxItem.tx_id.in_(tx_ids))
                   .order_by(TxItem.id).all()) if tx_ids else []
        vendors = {v.id: v for v in
                   db.query(Vendor).filter(Vendor.id.in_(vnd_ids)).all()} if vnd_ids else {}
        itm_map: Dict[int, List[TxItem]] = {}
        for it in all_its:
            itm_map.setdefault(it.tx_id, []).append(it)
        return [_tx_to_out(t, vendors.get(t.vendor_id, Vendor(name="")).name,
                           itm_map.get(t.id, [])) for t in txs]


@app.put("/api/tx/{tx_id}", response_model=TxOut)
def update_tx(tx_id: int, payload: TxIn, user: User = Depends(get_current_user)):
    tx_date          = payload.tx_date or datetime.now(timezone.utc)
    supply, vat, tot = _calc_amounts(payload.items, payload.vat_rate)
    with db_session() as db:
        t = db.query(Tx).filter(Tx.id == tx_id).first()
        if not t: raise HTTPException(404, "TX not found")
        vendor = db.query(Vendor).filter(Vendor.id == payload.vendor_id).first()
        if not vendor: raise HTTPException(400, "Invalid vendor_id")
        t.kind=payload.kind; t.tx_date=tx_date; t.vendor_id=payload.vendor_id
        t.description=payload.description; t.vat_rate=payload.vat_rate
        t.supply_amount=supply; t.vat_amount=vat; t.total_amount=tot
        if (payload.doc_no or "").strip(): t.doc_no = payload.doc_no.strip()
        db.query(TxItem).filter(TxItem.tx_id == tx_id).delete()
        for it in payload.items:
            db.add(TxItem(tx_id=tx_id, name=it.name,
                          spec=(it.spec or "").strip() or None,
                          qty=it.qty, unit_price=it.unit_price))
        db.commit(); db.refresh(t)
        return _tx_to_out(t, vendor.name, _items_of(db, t.id))


@app.delete("/api/tx/{tx_id}")
def delete_tx(tx_id: int, user: User = Depends(get_current_user)):
    with db_session() as db:
        t = db.query(Tx).filter(Tx.id == tx_id).first()
        if not t: raise HTTPException(404, "TX not found")
        db.delete(t); db.commit()
        return {"ok": True}


# ── 문서 다운로드 ─────────────────────────────────────────────
_DOC_FN = {DocType.견적서: "quote", DocType.발주서: "po", DocType.거래명세서: "statement"}


def _get_doc_objects(db: Session, tx_id: int, company_id: int):
    t = db.query(Tx).filter(Tx.id == tx_id).first()
    if not t: raise HTTPException(404, "TX not found")
    vendor = db.query(Vendor).filter(Vendor.id == t.vendor_id).first()
    if not vendor: raise HTTPException(404, "Vendor not found")
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company: raise HTTPException(404, "Company not found")
    return t, vendor, company, _items_of(db, t.id)


@app.get("/api/tx/{tx_id}/pdf")
def tx_pdf(tx_id: int,
           doc_type: DocType = Query(default=DocType.거래명세서),
           company_id: int = Query(default=1),
           user: User = Depends(get_current_user)):
    with db_session() as db:
        t, vendor, company, items = _get_doc_objects(db, tx_id, company_id)
        try:
            data = build_pdf(doc_type, t, vendor, company, items)
        except Exception as e:
            raise HTTPException(500, f"PDF 생성 실패: {e}")
    fname = f"{_DOC_FN.get(doc_type, 'doc')}_TX{tx_id}.pdf"
    return Response(content=data, media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@app.get("/api/tx/{tx_id}/excel")
def tx_excel(tx_id: int,
             doc_type: DocType = Query(default=DocType.거래명세서),
             company_id: int = Query(default=1),
             user: User = Depends(get_current_user)):
    with db_session() as db:
        t, vendor, company, items = _get_doc_objects(db, tx_id, company_id)
        try:
            data = build_excel(doc_type, t, vendor, company, items)
        except Exception as e:
            raise HTTPException(500, f"Excel 생성 실패: {e}")
    fname = f"{_DOC_FN.get(doc_type, 'doc')}_TX{tx_id}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ── Stats ─────────────────────────────────────────────────────
def _range_start(rk: str) -> Tuple[datetime, str]:
    now = datetime.now(timezone.utc); rk = (rk or "").strip().lower()
    if rk in ("7d","1w","week"):   return now - timedelta(days=7),   "최근 1주"
    if rk in ("1m","30d","month"): return now - timedelta(days=30),  "최근 1달"
    if rk in ("1y","365d","year"): return now - timedelta(days=365), "최근 1년"
    return now - timedelta(days=7), "최근 1주"


def _sum_rows(txs: List[Tx]) -> Dict[str, int]:
    return {"supply": sum(_safe_int(t.supply_amount) for t in txs),
            "vat":    sum(_safe_int(t.vat_amount)    for t in txs),
            "total":  sum(_safe_int(t.total_amount)  for t in txs)}


@app.get("/api/stats/summary")
def stats_summary(range: str = Query(default="7d"), user: User = Depends(get_current_user)):
    start, label = _range_start(range)
    with db_session() as db:
        rows  = db.query(Tx).filter(Tx.tx_date >= start).order_by(Tx.id.desc()).all()
        sales = [t for t in rows if t.kind == TxKind.매출]
        purch = [t for t in rows if t.kind == TxKind.매입]
        return {"range_key": range, "range_label": label,
                "start_utc": start.isoformat(), "count": len(rows),
                "sales": _sum_rows(sales), "purchase": _sum_rows(purch),
                "overall": _sum_rows(rows)}


@app.get("/api/stats/xlsx")
def stats_xlsx(range_key: str = Query(default="7d", alias="range"),
               user: User = Depends(get_current_user)):
    start, label = _range_start(range_key)
    with db_session() as db:
        txs   = (db.query(Tx).filter(Tx.tx_date >= start)
                 .order_by(Tx.tx_date.asc(), Tx.id.asc()).all())
        sales = [t for t in txs if t.kind == TxKind.매출]
        purch = [t for t in txs if t.kind == TxKind.매입]

        # ── 공통 스타일 헬퍼 ──────────────────────────────────
        FN   = "맑은 고딕"
        TK   = Side(border_style="thin", color="000000")
        FB   = Border(top=TK, bottom=TK, left=TK, right=TK)
        C_HDR  = PatternFill("solid", fgColor="FF2563EB")  # 진파랑
        C_SUB  = PatternFill("solid", fgColor="FFDBEAFE")  # 연파랑
        C_LBL  = PatternFill("solid", fgColor="FFEFF6FF")  # 아주연한파랑
        C_TOT  = PatternFill("solid", fgColor="FF1E40AF")  # 합계 짙은파랑
        C_ALT  = PatternFill("solid", fgColor="FFF0F9FF")  # 짝수행
        C_WHT  = PatternFill("solid", fgColor="FFFFFFFF")
        C_SALE = PatternFill("solid", fgColor="FFdbeafe")  # 매출 연파랑
        C_BUY  = PatternFill("solid", fgColor="FFfce7f3")  # 매입 연분홍

        def sc(ws, r, c, v="", bold=False, size=9, fc="1e293b",
               bg=None, align="left", numfmt=None):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font      = Font(bold=bold, size=size, color=fc, name=FN)
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border    = FB
            if bg:     cell.fill          = bg
            if numfmt: cell.number_format = numfmt
            return cell

        def apply_border(ws, max_row, max_col):
            for ri in range(1, max_row + 1):
                for ci in range(1, max_col + 1):
                    ws.cell(row=ri, column=ci).border = FB

        def set_col_widths(ws, widths):
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        # ══════════════════════════════════════════════════════
        #  시트 1: Summary
        # ══════════════════════════════════════════════════════
        wb   = Workbook()
        ws_s = wb.active
        ws_s.title = "요약"
        ws_s.page_setup.paperSize = ws_s.PAPERSIZE_A4
        ws_s.page_setup.fitToPage = True
        ws_s.page_setup.fitToWidth = 1
        set_col_widths(ws_s, [14, 20, 20, 20])

        now_str = _to_kst(datetime.now(timezone.utc)).strftime("%Y년 %m월 %d일 %H:%M")
        R = 1

        # 제목
        ws_s.merge_cells(start_row=R, start_column=1, end_row=R, end_column=4)
        sc(ws_s, R, 1, f"거래 통계  —  {label}", bold=True, size=16,
           fc="FFFFFF", bg=C_HDR, align="center")
        ws_s.row_dimensions[R].height = 36; R += 1

        # 기간/출력일
        ws_s.merge_cells(start_row=R, start_column=1, end_row=R, end_column=2)
        sc(ws_s, R, 1, f"기간: {label}", bold=True, size=9, bg=C_LBL)
        ws_s.merge_cells(start_row=R, start_column=3, end_row=R, end_column=4)
        sc(ws_s, R, 3, f"출력일: {now_str}", size=9, bg=C_LBL, align="right")
        ws_s.row_dimensions[R].height = 18; R += 1

        # 총 건수
        ws_s.merge_cells(start_row=R, start_column=1, end_row=R, end_column=4)
        sc(ws_s, R, 1, f"총 {len(txs)}건  (매출 {len(sales)}건 / 매입 {len(purch)}건)",
           size=9, bg=C_LBL)
        ws_s.row_dimensions[R].height = 16; R += 2

        # 합계 표 헤더
        ws_s.row_dimensions[R].height = 22
        for ci, hd in enumerate(["구분","공급가액","부가세","합계"], 1):
            sc(ws_s, R, ci, hd, bold=True, size=10, fc="FFFFFF",
               bg=C_HDR, align="center")
        R += 1

        s, p, o = _sum_rows(sales), _sum_rows(purch), _sum_rows(txs)
        for label_txt, d, is_tot, bg_row in [
            ("매출", s, False, C_SALE),
            ("매입", p, False, C_BUY),
            ("전  체", o, True,  C_TOT),
        ]:
            ws_s.row_dimensions[R].height = 20 if is_tot else 17
            fc = "FFFFFF" if is_tot else "1e293b"
            sz = 10 if is_tot else 9
            sc(ws_s, R, 1, label_txt, bold=is_tot, size=sz, fc=fc, bg=bg_row, align="center")
            sc(ws_s, R, 2, d["supply"], bold=is_tot, size=sz, fc=fc, bg=bg_row,
               align="right", numfmt="#,##0")
            sc(ws_s, R, 3, d["vat"],    bold=is_tot, size=sz, fc=fc, bg=bg_row,
               align="right", numfmt="#,##0")
            sc(ws_s, R, 4, d["total"],  bold=is_tot, size=sz, fc=fc, bg=bg_row,
               align="right", numfmt="#,##0")
            R += 1

        apply_border(ws_s, R - 1, 4)
        ws_s.print_area = f"A1:D{R - 1}"

        # ══════════════════════════════════════════════════════
        #  시트 2/3: 매출 / 매입 상세
        # ══════════════════════════════════════════════════════
        HDRS = ["ID","일시(KST)","거래처","문서번호","공급가","부가세","합계","메모"]
        COLS = [8, 20, 22, 24, 14, 14, 16, 30]
        ALIGNS = ["center","center","left","left","right","right","right","left"]
        NUMFMTS = [None, None, None, None, "#,##0", "#,##0", "#,##0", None]

        for sname, rows, hdr_bg in [("매출 상세", sales, C_SALE),
                                     ("매입 상세", purch, C_BUY)]:
            ws2 = wb.create_sheet(sname)
            ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
            ws2.page_setup.orientation = "landscape"
            ws2.page_setup.fitToPage = True
            ws2.page_setup.fitToWidth = 1
            set_col_widths(ws2, COLS)

            # 시트 제목
            ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
            sc(ws2, 1, 1, f"{sname}  —  {label}  ({len(rows)}건)",
               bold=True, size=14, fc="FFFFFF", bg=C_HDR, align="center")
            ws2.row_dimensions[1].height = 30

            # 헤더
            ws2.row_dimensions[2].height = 22
            for ci, (hd, al) in enumerate(zip(HDRS, ALIGNS), 1):
                sc(ws2, 2, ci, hd, bold=True, size=9.5, fc="FFFFFF",
                   bg=C_HDR, align="center")

            # 데이터
            for j, t in enumerate(rows, 1):
                R2 = j + 2
                ws2.row_dimensions[R2].height = 16
                bg = C_ALT if j % 2 == 0 else C_WHT
                vals = [
                    t.id,
                    _to_kst(t.tx_date).strftime("%Y-%m-%d %H:%M"),
                    t.vendor.name if t.vendor else "",
                    t.doc_no or "",
                    _safe_int(t.supply_amount),
                    _safe_int(t.vat_amount),
                    _safe_int(t.total_amount),
                    (t.description or "")[:40],
                ]
                for ci, (v, al, nf) in enumerate(zip(vals, ALIGNS, NUMFMTS), 1):
                    sc(ws2, R2, ci, v, size=9, bg=bg, align=al, numfmt=nf)

            # 합계 행
            if rows:
                tot_r = len(rows) + 3
                ws2.row_dimensions[tot_r].height = 20
                sr = _sum_rows(rows)
                ws2.merge_cells(start_row=tot_r, start_column=1, end_row=tot_r, end_column=4)
                sc(ws2, tot_r, 1, f"합계  ({len(rows)}건)", bold=True,
                   size=10, fc="FFFFFF", bg=C_TOT, align="center")
                for ci, key in [(5,"supply"),(6,"vat"),(7,"total")]:
                    sc(ws2, tot_r, ci, sr[key], bold=True, size=10,
                       fc="FFFFFF", bg=C_TOT, align="right", numfmt="#,##0")
                sc(ws2, tot_r, 8, "", bg=C_TOT)
                apply_border(ws2, tot_r, 8)
                ws2.print_area = f"A1:H{tot_r}"
            else:
                apply_border(ws2, 2, 8)

        bio = BytesIO()
        wb.save(bio)

    fname = f"ledger_stats_{range_key}.xlsx"
    return Response(content=bio.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})