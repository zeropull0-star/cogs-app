# api/main.py
# -*- coding: utf-8 -*-
from __future__ import annotations

import os
from datetime import datetime, timedelta
from enum import Enum
from io import BytesIO
from typing import List, Optional, Dict, Tuple
from zoneinfo import ZoneInfo

from fastapi import FastAPI, HTTPException, Query, Depends, status, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from fastapi.security import OAuth2PasswordBearer

from pydantic import BaseModel, Field, ConfigDict

from jose import JWTError, jwt
from passlib.context import CryptContext

from sqlalchemy import (
    create_engine,
    Column,
    Integer,
    String,
    DateTime,
    ForeignKey,
    Numeric,
    Enum as SAEnum,
    Text,
    Index,
    Boolean,
    text,
)
from sqlalchemy.orm import declarative_base, relationship, sessionmaker, Session

# ===== PDF (ReportLab) =====
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.lib.utils import ImageReader

# ===== Excel =====
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# 한글 폰트(CID)
pdfmetrics.registerFont(UnicodeCIDFont("HYGothic-Medium"))
pdfmetrics.registerFont(UnicodeCIDFont("HYSMyeongJo-Medium"))

W, H = A4

# ===== Config =====
DATABASE_URL = os.getenv("DATABASE_URL", "postgresql+psycopg://postgres:postgres@db:5432/ledger")

# assets
SEAL_PNG_PATH = (os.getenv("SEAL_PNG_PATH", "").strip() or None)
LOGO_PNG_PATH = (os.getenv("LOGO_PNG_PATH", "").strip() or None)  # 없으면 /app/assets/logo.png

COMPANY_NAME = os.getenv("COMPANY_NAME", "부원정보 주식회사")
COMPANY_BIZNO = os.getenv("COMPANY_BIZNO", "328-81-02550")
COMPANY_MANAGER = os.getenv("COMPANY_MANAGER", "김현수 부장")
COMPANY_PHONE = os.getenv("COMPANY_PHONE", "010-4236-8448")
COMPANY_ADDR = os.getenv("COMPANY_ADDR", "강원특별자치도 동해시 중앙로 173 2층")

DOCNO_PREFIX = (os.getenv("DOCNO_PREFIX", "BWIS") or "BWIS").strip()
DOCNO_DATE_FMT = (os.getenv("DOCNO_DATE_FMT", "%Y%m%d") or "%Y%m%d").strip()

# ===== Auth =====
SECRET_KEY = os.getenv("SECRET_KEY", "CHANGE_ME_TO_A_RANDOM_LONG_SECRET").strip()
ALGORITHM = (os.getenv("JWT_ALGORITHM", "HS256") or "HS256").strip()
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", "1440"))

ADMIN_USERNAME = (os.getenv("ADMIN_USERNAME", "admin") or "admin").strip()
ADMIN_PASSWORD = (os.getenv("ADMIN_PASSWORD", "admin1234") or "admin1234").strip()

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login")

# ===== DB =====
engine = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
Base = declarative_base()


class TxKind(str, Enum):
    매입 = "매입"
    매출 = "매출"


class DocType(str, Enum):
    견적서 = "견적서"
    발주서 = "발주서"
    거래명세서 = "거래명세서"


class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True)
    username = Column(String(200), nullable=False, unique=True, index=True)
    password_hash = Column(String(300), nullable=False)
    is_active = Column(Boolean, nullable=False, default=True)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)


class Vendor(Base):
    __tablename__ = "vendors"
    id = Column(Integer, primary_key=True)
    name = Column(String(200), nullable=False, index=True)
    biz_no = Column(String(64), nullable=True)
    ceo = Column(String(100), nullable=True)
    addr = Column(String(300), nullable=True)
    phone = Column(String(50), nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)

    txs = relationship("Tx", back_populates="vendor")


class Tx(Base):
    __tablename__ = "tx"

    id = Column(Integer, primary_key=True)

    kind = Column(SAEnum(TxKind, name="tx_kind"), nullable=False, index=True)
    tx_date = Column(DateTime, nullable=False, default=datetime.utcnow, index=True)

    vendor_id = Column(Integer, ForeignKey("vendors.id", ondelete="RESTRICT"), nullable=False)
    description = Column(Text, nullable=True)

    vat_rate = Column(Numeric(6, 4), nullable=False, default=0.10)

    supply_amount = Column(Numeric(18, 0), nullable=False, default=0)
    vat_amount = Column(Numeric(18, 0), nullable=False, default=0)
    total_amount = Column(Numeric(18, 0), nullable=False, default=0)

    doc_no = Column(String(64), nullable=True, index=True)

    vendor = relationship("Vendor", back_populates="txs")
    items = relationship(
        "TxItem",
        back_populates="tx",
        cascade="all, delete-orphan",
        order_by="TxItem.id",
    )


class TxItem(Base):
    __tablename__ = "tx_items"

    id = Column(Integer, primary_key=True)
    tx_id = Column(Integer, ForeignKey("tx.id", ondelete="CASCADE"), nullable=False, index=True)

    name = Column(String(200), nullable=False)
    spec = Column(String(200), nullable=True)  # 규격
    qty = Column(Numeric(18, 4), nullable=False, default=1)
    unit_price = Column(Numeric(18, 0), nullable=False, default=0)

    tx = relationship("Tx", back_populates="items")


Index("ix_tx_kind_date", Tx.kind, Tx.tx_date.desc())


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None) -> str:
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
    to_encode.update({"exp": expire, "iat": datetime.utcnow()})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)


def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)


def db_session() -> Session:
    return SessionLocal()


def init_db():
    Base.metadata.create_all(bind=engine)
    with SessionLocal() as db:
        u = db.query(User).filter(User.username == ADMIN_USERNAME).first()
        if not u:
            u = User(
                username=ADMIN_USERNAME,
                password_hash=get_password_hash(ADMIN_PASSWORD),
                is_active=True,
            )
            db.add(u)
            db.commit()


def ensure_schema():
    """
    기존 DB가 이미 만들어진 상태에서 spec 컬럼이 없으면 create_all()로는 추가가 안됨.
    그래서 startup 때 안전하게 ALTER TABLE로 보강.
    """
    try:
        with engine.begin() as conn:
            conn.execute(text("ALTER TABLE tx_items ADD COLUMN IF NOT EXISTS spec VARCHAR(200);"))
    except Exception:
        pass


# ===== Schemas =====
class TokenOut(BaseModel):
    access_token: str
    token_type: str = "bearer"
    expires_in_min: int


class UserOut(BaseModel):
    id: int
    username: str
    is_active: bool


class VendorIn(BaseModel):
    model_config = ConfigDict(extra="forbid")
    name: str
    biz_no: Optional[str] = None
    ceo: Optional[str] = None
    addr: Optional[str] = None
    phone: Optional[str] = None


class VendorOut(VendorIn):
    id: int
    created_at: datetime


class TxItemIn(BaseModel):
    model_config = ConfigDict(extra="forbid")
    name: str
    spec: Optional[str] = None
    qty: float = Field(ge=0)
    unit_price: float = Field(ge=0)


class TxIn(BaseModel):
    model_config = ConfigDict(extra="forbid")
    kind: TxKind
    tx_date: Optional[datetime] = None
    vendor_id: int
    description: Optional[str] = None
    vat_rate: float = Field(default=0.10, ge=0, le=1)
    doc_no: Optional[str] = None
    items: List[TxItemIn] = Field(default_factory=list)


class TxOut(BaseModel):
    id: int
    kind: TxKind
    tx_date: datetime
    vendor_id: int
    vendor_name: str
    description: Optional[str]
    vat_rate: float
    supply_amount: int
    vat_amount: int
    total_amount: int
    doc_no: Optional[str]
    items: List[TxItemIn]


# ===== Auth =====
def get_current_user(token: str = Depends(oauth2_scheme)) -> User:
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Unauthorized",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str | None = payload.get("sub")
        if not username:
            raise credentials_exception
    except JWTError:
        raise credentials_exception

    with db_session() as db:
        user = db.query(User).filter(User.username == username).first()
        if not user or not user.is_active:
            raise credentials_exception
        return user


# ===== Helpers =====
UTC = ZoneInfo("UTC")
KST = ZoneInfo("Asia/Seoul")


def _to_kst(dt: datetime) -> datetime:
    # DB에 naive로 저장된 경우가 많아서 "UTC로 저장되어 있다"로 보고 KST로 변환
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=UTC)
    return dt.astimezone(KST)


def _fmt_money(n: int) -> str:
    try:
        return f"{int(n):,}"
    except Exception:
        return "0"


def _safe_int(v) -> int:
    try:
        return int(round(float(v)))
    except Exception:
        return 0


def _extract_subject_from_memo(memo: Optional[str], max_len: int = 50) -> str:
    if not memo:
        return ""
    s = (memo or "").strip().replace("\r\n", "\n").replace("\r", "\n")
    first_line = (s.split("\n")[0] if s else "").strip()
    if not first_line:
        return ""
    if len(first_line) > max_len:
        return first_line[: max_len - 1] + "…"
    return first_line


def _calc_amounts(items: List[TxItemIn], vat_rate: float) -> Tuple[int, int, int]:
    supply = 0
    for it in items:
        supply += _safe_int(it.qty * it.unit_price)
    vat = _safe_int(supply * float(vat_rate))
    total = supply + vat
    return supply, vat, total


def _gen_doc_no(dt: datetime) -> str:
    return f"{DOCNO_PREFIX}-{dt.strftime(DOCNO_DATE_FMT)}"


def _draw_box(c: canvas.Canvas, x, y, w, h, stroke=colors.HexColor("#333333"), fill=None, radius=6):
    c.setStrokeColor(stroke)
    if fill:
        c.setFillColor(fill)
        c.roundRect(x, y, w, h, radius, stroke=1, fill=1)
    else:
        c.roundRect(x, y, w, h, radius, stroke=1, fill=0)


def _maybe_draw_image(c: canvas.Canvas, path: str, x: float, y: float, w: float, h: float):
    if not path:
        return
    if not os.path.exists(path):
        return
    try:
        img = ImageReader(path)
        c.drawImage(img, x, y, width=w, height=h, mask="auto", preserveAspectRatio=True, anchor="sw")
    except Exception:
        return


# ===== PDF =====
def _pdf_header(c: canvas.Canvas, title: str):
    margin = 20 * mm

    logo_path = LOGO_PNG_PATH or "/app/assets/logo.png"

    logo_w = 51 * mm
    logo_h = 27 * mm

    title_font = 26
    info_font = 10

    seal_path = SEAL_PNG_PATH
    seal_size = 23 * mm

    # 1) 로고
    logo_top_y = H - 10 * mm
    _maybe_draw_image(
        c,
        logo_path,
        x=margin,
        y=logo_top_y - logo_h + 6,
        w=logo_w,
        h=logo_h,
    )

    # 2) 타이틀(띄워쓰기) + 박스
    c.setFont("HYGothic-Medium", title_font)
    spaced_title = " ".join(title)

    title_center_y = H - 20 * mm

    pad_x = 5 * mm
    pad_y = 1 * mm

    text_w = pdfmetrics.stringWidth(spaced_title, "HYGothic-Medium", title_font)
    box_w = text_w + (pad_x * 2)
    box_h = (title_font * 0.85) * mm + (pad_y * 2)

    box_x = (W - box_w) / 2
    box_y = title_center_y - (box_h / 2)

    c.setLineWidth(1.2)
    c.setStrokeColor(colors.HexColor("#333333"))
    c.setFillColor(colors.white)
    c.roundRect(box_x, box_y, box_w, box_h, radius=6, stroke=1, fill=1)

    text_y = box_y + (box_h / 2) - 9
    c.setFillColor(colors.black)
    c.drawCentredString(W / 2, text_y, spaced_title)

    # 3) 회사정보(우측) - 라벨/값 열 정렬 + 주소 2줄
    c.setFont("HYGothic-Medium", info_font)
    c.setFillColor(colors.black)

    value_right = W - margin
    label_x = W - margin - 55 * mm

    row1_y = H - 20 * mm
    row_gap = 6 * mm

    addr1 = (COMPANY_ADDR or "").strip()
    addr2 = ""
    if "/" in addr1:
        p1, p2 = addr1.split("/", 1)
        addr1, addr2 = p1.strip(), p2.strip()
    else:
        parts = addr1.split()
        if len(parts) >= 2:
            addr1 = parts[0].strip()
            addr2 = " ".join(parts[1:]).strip()

    c.drawString(label_x, row1_y, "회사명")
    c.drawRightString(value_right, row1_y, COMPANY_NAME)

    c.drawString(label_x, row1_y - row_gap, "사업자등록번호")
    c.drawRightString(value_right, row1_y - row_gap, COMPANY_BIZNO)

    c.drawString(label_x, row1_y - (row_gap * 2), "담당")
    c.drawRightString(value_right, row1_y - (row_gap * 2), COMPANY_MANAGER)

    c.drawString(label_x, row1_y - (row_gap * 3), "연락처")
    c.drawRightString(value_right, row1_y - (row_gap * 3), COMPANY_PHONE)

    c.drawString(label_x, row1_y - (row_gap * 4), "주소")
    c.drawRightString(value_right, row1_y - (row_gap * 4), addr1)
    if addr2:
        c.drawRightString(value_right, row1_y - (row_gap * 5), addr2)

    # 4) 직인(마지막)
    seal_x = W - margin - seal_size + 13 * mm
    seal_y = H - (10 * mm) - seal_size
    _maybe_draw_image(c, seal_path, seal_x, seal_y, seal_size, seal_size)


def build_pdf(doc_type: DocType, tx: Tx, vendor: Vendor, items: List[TxItem]) -> bytes:
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)

    display_title = "명세서" if doc_type == DocType.거래명세서 else doc_type.value
    _pdf_header(c, display_title)

    margin = 20 * mm

    # 1) 거래처 박스
    top_y = H - 94 * mm
    box_h = 42 * mm
    _draw_box(c, margin, top_y, W - 2 * margin, box_h, stroke=colors.HexColor("#555555"))

    c.setFont("HYGothic-Medium", 10)
    c.setFillColor(colors.black)

    left_label_x = margin + 6 * mm
    left_value_x = margin + 18 * mm

    right_label_x = W - margin - 53 * mm
    right_value_x = W - margin - 6 * mm

    row1_y = top_y + box_h - 10 * mm
    row_gap = 8 * mm

    # 좌측
    c.drawString(left_label_x, row1_y, "거래처")
    c.drawString(left_value_x, row1_y, f": {vendor.name}")

    c.drawString(left_label_x, row1_y - row_gap, "대표자")
    c.drawString(left_value_x, row1_y - row_gap, f": {vendor.ceo or ''}")

    c.drawString(left_label_x, row1_y - (row_gap * 2), "연락처")
    c.drawString(left_value_x, row1_y - (row_gap * 2), f": {vendor.phone or ''}")

    c.drawString(left_label_x, row1_y - (row_gap * 3), "주  소")
    c.drawString(left_value_x, row1_y - (row_gap * 3), f": {vendor.addr or ''}")

    # 우측
    c.drawString(right_label_x, row1_y, "문서번호")
    c.drawRightString(right_value_x, row1_y, tx.doc_no or "")

    c.drawString(right_label_x, row1_y - row_gap, "일자")
    c.drawRightString(right_value_x, row1_y - row_gap, _to_kst(tx.tx_date).strftime("%Y-%m-%d"))

    # 2) 합계 박스 위치(테이블 계산용)
    sum_y = 38 * mm
    sum_h = 34 * mm

    # 3) 건명 박스
    subject_h = 10 * mm
    gap_vendor_subject = 3 * mm
    gap_subject_table = 6 * mm

    subject_y = top_y - gap_vendor_subject - subject_h
    _draw_box(c, margin, subject_y, W - 2 * margin, subject_h, stroke=colors.HexColor("#555555"))

    subject_label_x = margin + 6 * mm
    subject_value_x = margin + 18 * mm
    subject_text = _extract_subject_from_memo(tx.description, max_len=60)

    c.drawString(subject_label_x, subject_y + subject_h - 6 * mm, "건명")
    c.drawString(subject_value_x, subject_y + subject_h - 6 * mm, f": {subject_text}")

    # 4) 품목 테이블 박스(자동 계산)
    gap_table_sum = 12 * mm

    table_top = subject_y - gap_subject_table
    table_bottom = sum_y + sum_h + gap_table_sum

    table_y = table_bottom
    table_h = table_top - table_bottom

    if table_h < 50 * mm:
        table_h = 50 * mm
        table_y = table_top - table_h

    _draw_box(c, margin, table_y, W - 2 * margin, table_h, stroke=colors.HexColor("#555555"))

    x_name = margin + 6 * mm
    x_spec = margin + 60 * mm
    x_qty = margin + 105 * mm
    x_unit = margin + 128 * mm
    x_amtR = W - margin - 6 * mm

    c.drawString(x_name, table_y + table_h - 10 * mm, "품목")
    c.drawString(x_spec, table_y + table_h - 10 * mm, "규격")
    c.drawString(x_qty, table_y + table_h - 10 * mm, "수량")
    c.drawString(x_unit, table_y + table_h - 10 * mm, "단가")
    c.drawRightString(x_amtR, table_y + table_h - 10 * mm, "금액")

    c.setStrokeColor(colors.HexColor("#777777"))
    c.setLineWidth(0.6)
    c.line(margin + 4 * mm, table_y + table_h - 14 * mm, W - margin - 4 * mm, table_y + table_h - 14 * mm)

    row_y = table_y + table_h - 24 * mm
    row_h = 10 * mm

    for it in items[:8]:
        amount = _safe_int(float(it.qty) * float(it.unit_price))
        c.drawString(x_name, row_y, it.name or "")
        c.drawString(x_spec, row_y, it.spec or "")
        c.drawString(x_qty, row_y, f"{float(it.qty):g}")
        c.drawString(x_unit, row_y, _fmt_money(_safe_int(it.unit_price)))
        c.drawRightString(x_amtR, row_y, _fmt_money(amount))
        row_y -= row_h

    # 5) 합계 박스
    _draw_box(c, margin, sum_y, W - 2 * margin, sum_h, stroke=colors.HexColor("#555555"))

    label_x = margin + 10 * mm
    value_x = W - margin - 10 * mm

    row_top = sum_y + sum_h - 10 * mm
    row_gap = 8 * mm

    c.setFont("HYGothic-Medium", 11)
    c.setFillColor(colors.black)

    c.drawString(label_x, row_top, "공급가액")
    c.drawRightString(value_x, row_top, f"{_fmt_money(_safe_int(tx.supply_amount))} 원")

    c.drawString(label_x, row_top - row_gap, "부가세")
    c.drawRightString(value_x, row_top - row_gap, f"{_fmt_money(_safe_int(tx.vat_amount))} 원")

    c.setStrokeColor(colors.HexColor("#999999"))
    c.setLineWidth(0.8)
    c.line(margin + 8 * mm, row_top - row_gap - 4 * mm, W - margin - 8 * mm, row_top - row_gap - 4 * mm)

    total_y = row_top - (row_gap * 2) - 2 * mm
    c.setFont("HYGothic-Medium", 14)

    c.drawString(label_x, total_y, "합계")
    c.drawRightString(value_x, total_y, f"{_fmt_money(_safe_int(tx.total_amount))} 원")

    c.showPage()
    c.save()
    return buf.getvalue()


# ===== App =====
app = FastAPI(title="Ledger API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_headers=["*"],
    allow_methods=["*"],
)


@app.on_event("startup")
def _startup():
    init_db()
    ensure_schema()


@app.get("/health")
def health():
    return {"ok": True}


# ===== Auth =====
@app.post("/api/auth/login", response_model=TokenOut)
def login(username: str = Form(...), password: str = Form(...)):
    with db_session() as db:
        u = db.query(User).filter(User.username == username).first()
        if not u or not verify_password(password, u.password_hash) or not u.is_active:
            raise HTTPException(status_code=400, detail="Invalid credentials")
        token = create_access_token({"sub": u.username})
        return TokenOut(access_token=token, token_type="bearer", expires_in_min=ACCESS_TOKEN_EXPIRE_MINUTES)


@app.get("/api/auth/me", response_model=UserOut)
def me(user: User = Depends(get_current_user)):
    return UserOut(id=user.id, username=user.username, is_active=user.is_active)


# ===== Vendors =====
@app.post("/api/vendors", response_model=VendorOut)
def create_vendor(payload: VendorIn, user: User = Depends(get_current_user)):
    with db_session() as db:
        v = Vendor(**payload.model_dump())
        db.add(v)
        db.commit()
        db.refresh(v)
        return VendorOut(
            id=v.id,
            name=v.name,
            biz_no=v.biz_no,
            ceo=v.ceo,
            addr=v.addr,
            phone=v.phone,
            created_at=v.created_at,
        )


@app.get("/api/vendors", response_model=List[VendorOut])
def list_vendors(q: Optional[str] = Query(default=None), user: User = Depends(get_current_user)):
    with db_session() as db:
        qs = db.query(Vendor)
        if q:
            qs = qs.filter(Vendor.name.ilike(f"%{q}%"))
        rows = qs.order_by(Vendor.id.desc()).all()
        return [
            VendorOut(
                id=v.id,
                name=v.name,
                biz_no=v.biz_no,
                ceo=v.ceo,
                addr=v.addr,
                phone=v.phone,
                created_at=v.created_at,
            )
            for v in rows
        ]


@app.put("/api/vendors/{vendor_id}", response_model=VendorOut)
def update_vendor(vendor_id: int, payload: VendorIn, user: User = Depends(get_current_user)):
    with db_session() as db:
        v = db.query(Vendor).filter(Vendor.id == vendor_id).first()
        if not v:
            raise HTTPException(status_code=404, detail="Vendor not found")
        for k, val in payload.model_dump().items():
            setattr(v, k, val)
        db.commit()
        db.refresh(v)
        return VendorOut(
            id=v.id,
            name=v.name,
            biz_no=v.biz_no,
            ceo=v.ceo,
            addr=v.addr,
            phone=v.phone,
            created_at=v.created_at,
        )


@app.delete("/api/vendors/{vendor_id}")
def delete_vendor(vendor_id: int, user: User = Depends(get_current_user)):
    with db_session() as db:
        v = db.query(Vendor).filter(Vendor.id == vendor_id).first()
        if not v:
            raise HTTPException(status_code=404, detail="Vendor not found")
        db.delete(v)
        db.commit()
        return {"ok": True}


# ===== TX =====
@app.post("/api/tx", response_model=TxOut)
def create_tx(payload: TxIn, user: User = Depends(get_current_user)):
    tx_date = payload.tx_date or datetime.utcnow()
    supply, vat, total = _calc_amounts(payload.items, payload.vat_rate)

    with db_session() as db:
        vendor = db.query(Vendor).filter(Vendor.id == payload.vendor_id).first()
        if not vendor:
            raise HTTPException(status_code=400, detail="Invalid vendor_id")

        tx = Tx(
            kind=payload.kind,
            tx_date=tx_date,
            vendor_id=payload.vendor_id,
            description=payload.description,
            vat_rate=payload.vat_rate,
            supply_amount=supply,
            vat_amount=vat,
            total_amount=total,
            doc_no=(payload.doc_no or "").strip() or None,
        )
        db.add(tx)
        db.flush()

        if not tx.doc_no:
            base = _gen_doc_no(tx_date)
            tx.doc_no = f"{base}-{tx.id:04d}"

        for it in payload.items:
            db.add(
                TxItem(
                    tx_id=tx.id,
                    name=it.name,
                    spec=(it.spec or "").strip() or None,
                    qty=it.qty,
                    unit_price=it.unit_price,
                )
            )

        db.commit()
        db.refresh(tx)

        items = db.query(TxItem).filter(TxItem.tx_id == tx.id).order_by(TxItem.id.asc()).all()

        return TxOut(
            id=tx.id,
            kind=tx.kind,
            tx_date=_to_kst(tx.tx_date),  # ✅ 장부 화면 시간(KST)로 내려줌
            vendor_id=tx.vendor_id,
            vendor_name=vendor.name,
            description=tx.description,
            vat_rate=float(tx.vat_rate),
            supply_amount=_safe_int(tx.supply_amount),
            vat_amount=_safe_int(tx.vat_amount),
            total_amount=_safe_int(tx.total_amount),
            doc_no=tx.doc_no,
            items=[TxItemIn(name=i.name, spec=i.spec, qty=float(i.qty), unit_price=float(i.unit_price)) for i in items],
        )


@app.get("/api/tx", response_model=List[TxOut])
def list_tx(
    kind: Optional[TxKind] = Query(default=None),
    vendor_id: Optional[int] = Query(default=None),
    limit: int = Query(default=200, ge=1, le=500),
    user: User = Depends(get_current_user),
):
    with db_session() as db:
        qs = db.query(Tx).order_by(Tx.id.desc())
        if kind:
            qs = qs.filter(Tx.kind == kind)
        if vendor_id:
            qs = qs.filter(Tx.vendor_id == vendor_id)

        txs = qs.limit(limit).all()

        vendor_ids = list({t.vendor_id for t in txs})
        vendors = {v.id: v for v in db.query(Vendor).filter(Vendor.id.in_(vendor_ids)).all()} if vendor_ids else {}

        out: List[TxOut] = []
        for t in txs:
            items = db.query(TxItem).filter(TxItem.tx_id == t.id).order_by(TxItem.id.asc()).all()
            v = vendors.get(t.vendor_id)
            out.append(
                TxOut(
                    id=t.id,
                    kind=t.kind,
                    tx_date=_to_kst(t.tx_date),  # ✅ 장부 화면 시간(KST)로 내려줌
                    vendor_id=t.vendor_id,
                    vendor_name=v.name if v else "",
                    description=t.description,
                    vat_rate=float(t.vat_rate),
                    supply_amount=_safe_int(t.supply_amount),
                    vat_amount=_safe_int(t.vat_amount),
                    total_amount=_safe_int(t.total_amount),
                    doc_no=t.doc_no,
                    items=[TxItemIn(name=i.name, spec=i.spec, qty=float(i.qty), unit_price=float(i.unit_price)) for i in items],
                )
            )
        return out


@app.put("/api/tx/{tx_id}", response_model=TxOut)
def update_tx(tx_id: int, payload: TxIn, user: User = Depends(get_current_user)):
    tx_date = payload.tx_date or datetime.utcnow()
    supply, vat, total = _calc_amounts(payload.items, payload.vat_rate)

    with db_session() as db:
        tx = db.query(Tx).filter(Tx.id == tx_id).first()
        if not tx:
            raise HTTPException(status_code=404, detail="TX not found")

        vendor = db.query(Vendor).filter(Vendor.id == payload.vendor_id).first()
        if not vendor:
            raise HTTPException(status_code=400, detail="Invalid vendor_id")

        tx.kind = payload.kind
        tx.tx_date = tx_date
        tx.vendor_id = payload.vendor_id
        tx.description = payload.description
        tx.vat_rate = payload.vat_rate
        tx.supply_amount = supply
        tx.vat_amount = vat
        tx.total_amount = total

        incoming_doc = (payload.doc_no or "").strip()
        if incoming_doc:
            tx.doc_no = incoming_doc

        db.query(TxItem).filter(TxItem.tx_id == tx_id).delete()
        for it in payload.items:
            db.add(
                TxItem(
                    tx_id=tx_id,
                    name=it.name,
                    spec=(it.spec or "").strip() or None,
                    qty=it.qty,
                    unit_price=it.unit_price,
                )
            )

        db.commit()
        db.refresh(tx)

        items = db.query(TxItem).filter(TxItem.tx_id == tx.id).order_by(TxItem.id.asc()).all()

        return TxOut(
            id=tx.id,
            kind=tx.kind,
            tx_date=_to_kst(tx.tx_date),  # ✅ 장부 화면 시간(KST)로 내려줌
            vendor_id=tx.vendor_id,
            vendor_name=vendor.name,
            description=tx.description,
            vat_rate=float(tx.vat_rate),
            supply_amount=_safe_int(tx.supply_amount),
            vat_amount=_safe_int(tx.vat_amount),
            total_amount=_safe_int(tx.total_amount),
            doc_no=tx.doc_no,
            items=[TxItemIn(name=i.name, spec=i.spec, qty=float(i.qty), unit_price=float(i.unit_price)) for i in items],
        )


@app.delete("/api/tx/{tx_id}")
def delete_tx(tx_id: int, user: User = Depends(get_current_user)):
    with db_session() as db:
        t = db.query(Tx).filter(Tx.id == tx_id).first()
        if not t:
            raise HTTPException(status_code=404, detail="TX not found")
        db.delete(t)
        db.commit()
        return {"ok": True}


DOC_FILENAME = {
    DocType.견적서: "quote",
    DocType.발주서: "po",
    DocType.거래명세서: "statement",
}


@app.get("/api/tx/{tx_id}/pdf")
def tx_pdf(
    tx_id: int,
    doc_type: DocType = Query(default=DocType.거래명세서),
    user: User = Depends(get_current_user),
):
    with db_session() as db:
        tx = db.query(Tx).filter(Tx.id == tx_id).first()
        if not tx:
            raise HTTPException(status_code=404, detail="TX not found")
        vendor = db.query(Vendor).filter(Vendor.id == tx.vendor_id).first()
        if not vendor:
            raise HTTPException(status_code=404, detail="Vendor not found")
        items = db.query(TxItem).filter(TxItem.tx_id == tx.id).order_by(TxItem.id.asc()).all()

        try:
            pdf_bytes = build_pdf(doc_type, tx, vendor, items)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"PDF build failed: {type(e).__name__}: {e}")

        prefix = DOC_FILENAME.get(doc_type, "doc")
        safe_name = f"{prefix}_TX{tx.id}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
        )


# ===== Stats / Backup =====
def _range_to_start(range_key: str) -> Tuple[datetime, str]:
    now = datetime.utcnow()
    rk = (range_key or "").strip().lower()

    if rk in ("7d", "1w", "week"):
        return now - timedelta(days=7), "최근 1주"
    if rk in ("1m", "30d", "month"):
        return now - timedelta(days=30), "최근 1달"
    if rk in ("1y", "365d", "year"):
        return now - timedelta(days=365), "최근 1년"

    return now - timedelta(days=7), "최근 1주"


def _sum_rows(txs: List[Tx]) -> Dict[str, int]:
    supply = sum(_safe_int(t.supply_amount) for t in txs)
    vat = sum(_safe_int(t.vat_amount) for t in txs)
    total = sum(_safe_int(t.total_amount) for t in txs)
    return {"supply": supply, "vat": vat, "total": total}


@app.get("/api/stats/summary")
def stats_summary(
    range: str = Query(default="7d"),
    user: User = Depends(get_current_user),
):
    start_dt, label = _range_to_start(range)

    with db_session() as db:
        rows = (
            db.query(Tx)
            .filter(Tx.tx_date >= start_dt)
            .order_by(Tx.id.desc())
            .all()
        )

        sales = [t for t in rows if t.kind == TxKind.매출]
        purchase = [t for t in rows if t.kind == TxKind.매입]

        out = {
            "range_key": range,
            "range_label": label,
            "start_utc": start_dt.isoformat(),
            "count": len(rows),
            "sales": _sum_rows(sales),
            "purchase": _sum_rows(purchase),
            "overall": _sum_rows(rows),
        }
        return out


@app.get("/api/stats/xlsx")
def stats_xlsx(
    range_key: str = Query(default="7d", alias="range"),
    user: User = Depends(get_current_user),
):
    start_dt, label = _range_to_start(range_key)

    with db_session() as db:
        txs: List[Tx] = (
            db.query(Tx)
            .filter(Tx.tx_date >= start_dt)
            .order_by(Tx.tx_date.asc(), Tx.id.asc())
            .all()
        )

        sales = [t for t in txs if t.kind == TxKind.매출]
        purchase = [t for t in txs if t.kind == TxKind.매입]

        wb = Workbook()

        # Summary
        ws_sum = wb.active
        ws_sum.title = "Summary"

        ws_sum.append(["기간", label])
        ws_sum.append(["시작(KST)", _to_kst(start_dt).strftime("%Y-%m-%d %H:%M:%S")])
        ws_sum.append(["건수", len(txs)])

        s_sum = _sum_rows(sales)
        p_sum = _sum_rows(purchase)
        o_sum = _sum_rows(txs)

        ws_sum.append([])
        ws_sum.append(["구분", "공급가", "부가세", "합계"])
        ws_sum.append(["매출", s_sum["supply"], s_sum["vat"], s_sum["total"]])
        ws_sum.append(["매입", p_sum["supply"], p_sum["vat"], p_sum["total"]])
        ws_sum.append(["전체", o_sum["supply"], o_sum["vat"], o_sum["total"]])

        for col in range(1, 5):
            ws_sum.column_dimensions[get_column_letter(col)].width = 18

        # 매출
        ws_sales = wb.create_sheet("매출")
        ws_sales.append(["TX_ID", "일시(KST)", "거래처", "문서번호", "공급가", "부가세", "합계", "메모"])
        for t in sales:
            kst_time = _to_kst(t.tx_date)
            ws_sales.append([
                t.id,
                kst_time.strftime("%Y-%m-%d %H:%M:%S"),
                t.vendor.name if t.vendor else "",
                t.doc_no or "",
                _safe_int(t.supply_amount),
                _safe_int(t.vat_amount),
                _safe_int(t.total_amount),
                t.description or "",
            ])

        # 매입
        ws_purchase = wb.create_sheet("매입")
        ws_purchase.append(["TX_ID", "일시(KST)", "거래처", "문서번호", "공급가", "부가세", "합계", "메모"])
        for t in purchase:
            kst_time = _to_kst(t.tx_date)
            ws_purchase.append([
                t.id,
                kst_time.strftime("%Y-%m-%d %H:%M:%S"),
                t.vendor.name if t.vendor else "",
                t.doc_no or "",
                _safe_int(t.supply_amount),
                _safe_int(t.vat_amount),
                _safe_int(t.total_amount),
                t.description or "",
            ])

        widths = [10, 20, 26, 22, 14, 14, 14, 30]
        for ws in [ws_sales, ws_purchase]:
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

        bio = BytesIO()
        wb.save(bio)
        xlsx_bytes = bio.getvalue()

    safe_name = f"ledger_stats_{range_key}.xlsx"
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
    )